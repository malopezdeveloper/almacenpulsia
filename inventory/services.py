import hashlib,io,json,re,unicodedata
from datetime import date,datetime
from decimal import Decimal
from openpyxl import load_workbook,Workbook
from django.db import transaction
from django.utils import timezone
from django.utils.text import slugify
from .models import InventoryTable,InventoryField,InventoryRecord,RecordMovement,Incident,ImportJob,AuditLog

def plain(value):
 if isinstance(value,(datetime,date)): return value.isoformat()
 if isinstance(value,Decimal): return float(value)
 return value

def normalized(value):
 text=unicodedata.normalize("NFKD",str(value or "")).encode("ascii","ignore").decode().lower()
 return re.sub(r"[^a-z0-9]+"," ",text).strip()

def unique_slug(value,used):
 base=(slugify(value).replace("-","_") or "campo")[:150]; key=base; n=2
 while key in used: key=f"{base}_{n}"; n+=1
 used.add(key); return key

def infer_type(header,values):
 name=normalized(header); present=[v for v in values if v not in (None,"")]
 if "fecha" in name: return "date"
 if present and all(isinstance(v,(int,float)) and not isinstance(v,bool) for v in present[:30]): return "number"
 if present and all(normalized(v) in {"si","no"} for v in present[:30]): return "bool"
 return "text"

def infer_id_pattern(values,fallback):
 patterns={}
 for value in values:
  match=re.match(r"^(.*?)(\d+)$",str(value or "").strip())
  if match:
   prefix,digits=match.groups(); bucket=patterns.setdefault(prefix,{"count":0,"width":0,"maximum":0}); bucket["count"]+=1; bucket["width"]=max(bucket["width"],len(digits)); bucket["maximum"]=max(bucket["maximum"],int(digits))
 if patterns:
  prefix,info=max(patterns.items(),key=lambda pair:(pair[1]["count"],pair[1]["maximum"])); return prefix,info["width"],info["maximum"]+1
 return (fallback.upper()[:8]+"-",4,1)

@transaction.atomic
def import_excel(upload,user):
 raw=upload.read(); digest=hashlib.sha256(raw).hexdigest(); job=ImportJob.objects.create(file_name=upload.name,fingerprint=digest,created_by=user)
 wb=load_workbook(io.BytesIO(raw),data_only=True,read_only=True)
 for sheet_position,ws in enumerate(wb.worksheets,1):
  all_rows=list(ws.iter_rows(values_only=True)); headers=[str(x).strip() if x is not None else "" for x in (all_rows[0] if all_rows else [])]
  table,created=InventoryTable.objects.get_or_create(name=ws.title,defaults={"slug":slugify(ws.title) or f"tabla-{sheet_position}","position":sheet_position,"created_by":user})
  if not created and table.position!=sheet_position: table.position=sheet_position; table.save(update_fields=["position"])
  nonempty=[i for i,h in enumerate(headers) if h]
  if not nonempty:
   Incident.objects.create(title=f"Hoja sin encabezados: {ws.title}",details="No se pudo definir la tabla.",kind="missing_headers",severity="error",source_file=upload.name,source_sheet=ws.title); job.rows_incident+=1; continue
  primary_index=nonempty[0]; table.id_header=headers[primary_index]; table.save(update_fields=["id_header"])
  prefix,width,next_number=infer_id_pattern([row[primary_index] if primary_index<len(row) else None for row in all_rows[1:]],slugify(ws.title).replace("-","") or "OBJ")
  if created or not table.id_prefix or next_number>table.next_number: table.id_prefix=prefix; table.id_width=width; table.next_number=next_number; table.save(update_fields=["id_prefix","id_width","next_number"])
  used=set(); field_map=[]
  seen_names=set()
  for position,index in enumerate(nonempty):
   header=headers[index]; key=unique_slug(header,used); name_key=normalized(header)
   samples=[row[index] if index<len(row) else None for row in all_rows[1:31]]
   is_primary=index==primary_index; is_sn=name_key in {"sn","sn destino"} or name_key.startswith("sn destino "); is_tech="tecnico" in name_key
   field,_=InventoryField.objects.update_or_create(table=table,key=key,defaults={"name":header,"position":position,"field_type":infer_type(header,samples),"is_primary":is_primary,"is_destination_sn":is_sn,"is_technician":is_tech})
   field_map.append((index,field))
   if name_key in seen_names:
    Incident.objects.create(title=f"Encabezado duplicado: {header}",details="Se conservó como un campo independiente.",kind="duplicate_header",source_file=upload.name,source_sheet=ws.title,payload={"field":header,"key":key}); job.rows_incident+=1
   seen_names.add(name_key)
  valid_keys={field.key for _,field in field_map}; table.inventory_fields.exclude(key__in=valid_keys).delete()
  for row_no,row in enumerate(all_rows[1:],2):
   if not any(v not in (None,"") for v in row): continue
   job.rows_total+=1; internal=str(row[primary_index] or "").strip() if primary_index<len(row) else ""
   payload={field.name:plain(row[index]) for index,field in field_map if index<len(row) and row[index] not in (None,"")}
   unnamed={str(i+1):plain(v) for i,v in enumerate(row) if i<len(headers) and not headers[i] and v not in (None,"")}
   if unnamed:
    Incident.objects.create(title="Datos en columnas sin nombre",details="Los valores no se importaron porque la columna no tiene encabezado.",kind="unnamed_column",source_file=upload.name,source_sheet=ws.title,source_row=row_no,payload=unnamed); job.rows_incident+=1
   if not internal:
    Incident.objects.create(title="Fila sin ID interno",details="La fila contiene datos pero no una clave primaria.",kind="missing_id",severity="error",source_file=upload.name,source_sheet=ws.title,source_row=row_no,payload=payload); job.rows_incident+=1; continue
   if InventoryRecord.objects.filter(internal_id__iexact=internal).exists():
    incident_payload=dict(payload); incident_payload["__duplicate_internal_id"]=internal; incident_payload["__inventory_table_pk"]=table.pk
    Incident.objects.create(title=f"ID duplicado en {ws.title}: {internal}",details="El registro quedó aislado y no se sobrescribió ningún objeto. Debe resolverse físicamente desde Incidencias.",kind="duplicate_id",severity="error",source_file=upload.name,source_sheet=ws.title,source_row=row_no,payload=incident_payload); job.rows_incident+=1; continue
   data={field.key:plain(row[index]) if index<len(row) and row[index] is not None else "" for index,field in field_map if not field.is_primary}
   sn=next((data.get(field.key) for _,field in field_map if field.is_destination_sn and data.get(field.key)),"")
   tech=next((data.get(field.key) for _,field in field_map if field.is_technician and data.get(field.key)),"")
   record=InventoryRecord.objects.create(table=table,internal_id=internal,data=data,current_sn=str(sn or ""),current_technician=str(tech or ""),status="assigned" if sn else "available",created_by=user)
   if sn or tech: RecordMovement.objects.create(record=record,movement_type="entry",technician_name=str(tech or ""),destination_sn=str(sn or ""),reason=f"Importado desde {upload.name}",registered_by=user)
   job.rows_imported+=1
 job.status="completed_with_incidents" if job.rows_incident else "completed"; job.save(); AuditLog.objects.create(user=user,action="excel_import",object_type="ImportJob",object_id=str(job.pk),details={"tables":len(wb.sheetnames),"imported":job.rows_imported,"incidents":job.rows_incident}); return job

def export_excel():
 wb=Workbook(); wb.remove(wb.active)
 for table in InventoryTable.objects.filter(active=True).prefetch_related("inventory_fields").order_by("position","name"):
  ws=wb.create_sheet(table.name[:31]); fields=list(table.inventory_fields.all()); ws.append([f.name for f in fields])
  for record in table.records.order_by("internal_id"):
   ws.append([record.internal_id if f.is_primary else record.data.get(f.key,"") for f in fields])
  ws.freeze_panes="A2"; ws.auto_filter.ref=ws.dimensions
 if not wb.sheetnames: wb.create_sheet("Inventario vacío").append(["Sin tablas"])
 out=io.BytesIO(); wb.save(out); return out.getvalue()
