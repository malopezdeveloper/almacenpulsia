import hashlib,json,os,re,secrets,sqlite3,string,tempfile,threading,time
from datetime import date,datetime,timedelta
from functools import wraps
from pathlib import Path
from django.conf import settings
from django.apps import apps
from django.contrib import messages
from django.contrib.auth import authenticate,get_user_model,login,logout
from django.contrib.auth.decorators import login_required
from django.core.paginator import Paginator
from django.core.management import call_command
from django.db import IntegrityError,connection,connections,transaction
from django.db.models import Q,Count,Sum
from django.http import FileResponse,HttpResponse,HttpResponseForbidden,JsonResponse
from django.shortcuts import get_object_or_404,redirect,render
from django.urls import reverse
from django.utils import timezone
from django.db.models.functions import TruncHour,TruncDay,TruncWeek,TruncMonth
from django.utils.http import url_has_allowed_host_and_scheme
from django.utils.text import slugify
from django.views.decorators.csrf import ensure_csrf_cookie
from .forms import PasswordChangeRequiredForm,InventoryTableForm,InventoryFieldForm,DynamicRecordForm,RecordAssignmentForm,RecordScrapForm,ReservationForm,ExistingLabelReprintForm,LabelSequenceForm,LoanRequestForm,LoanItemForm,ClientBatchSheetForm,ClientBatchFieldForm,ClientBatchRowForm,DuplicateIncidentResolutionForm
from .models import InventoryTable,InventoryField,InventoryRecord,RecordMovement,Incident,AuditLog,ImportJob,UserProfile,AccessUpgradeRequest,BackupSchedule,BackupDiskConfig,SecurityAccessPolicy,SecurityAccessEvent,ActiveSecuritySession,Reservation,ReservationView,ChatMessage,LabelPrintJob,Loan,LoanItem,LoanRequest,ClientBatchSheet,ClientBatchField,ClientBatchRow,ClientBatchChange,ServiceAccess,IPBan,NetworkReservationRequest,ProductionModelMySQLSource,ProductionModel,ProductionModelExclusion,ProductionProcessor,ProductionZone,ProductionEntry
from .printing import print_identifier
from .networking import request_current_ip_reservation, current_network
from .services import import_excel,export_excel
from .security import access_window_state, close_security_session, get_policy, login_allowed, register_client_fingerprint, register_login_session, seconds_until_logout_for_user
from .external_mysql import encrypt_password, fetch_models, normalize_model, test_source
from .storage_admin import request_storage_admin
from .db_utils import create_sqlite_snapshot, DeleteOnCloseFile
from .ip_utils import is_protected_local_ip

RESERVED_NAMES={"admin","administrator","administrador","root","system","sistema","almacen","_operador_local"}

# La búsqueda global recorre las tablas funcionales de la aplicación. Se excluyen
# tablas internas de Django y campos sensibles que nunca deben mostrarse.
GLOBAL_SEARCH_EXCLUDED_TABLES={
 "django_migrations","django_session","django_content_type","auth_permission",
 "auth_group_permissions","auth_user_groups","auth_user_user_permissions",
}
GLOBAL_SEARCH_HIDDEN_COLUMNS={"password","session_data","bootstrap_token_hash"}

def _global_table_label(table_name):
 try:
  for model in apps.get_models():
   if model._meta.db_table==table_name:
    return str(model._meta.verbose_name_plural or model._meta.verbose_name).capitalize()
 except Exception:
  pass
 return table_name

def global_database_search(query):
 """Busca una cadena, sin distinguir mayúsculas/minúsculas, en todas las tablas
 funcionales de la base de datos. Devuelve la fila y los campos que coincidieron.
 La consulta es de solo lectura y nunca expone columnas sensibles.
 """
 query=(query or "").strip()
 if not query: return []
 needle=query.casefold(); results=[]; qn=connection.ops.quote_name
 with connection.cursor() as cursor:
  table_names=connection.introspection.table_names(cursor)
  for table_name in sorted(table_names):
   if table_name in GLOBAL_SEARCH_EXCLUDED_TABLES: continue
   # Buscamos tablas de modelos instalados. Esto incluye inventory_* y auth_user,
   # pero evita tablas auxiliares ajenas a los modelos de la aplicación.
   model=None
   for candidate in apps.get_models():
    if candidate._meta.db_table==table_name:
     model=candidate; break
   if model is None: continue
   try:
    description=connection.introspection.get_table_description(cursor,table_name)
    columns=[c.name for c in description if c.name not in GLOBAL_SEARCH_HIDDEN_COLUMNS]
    if not columns: continue
    # Se leen las columnas visibles y se compara en Python para que el comportamiento
    # sea idéntico con SQLite/PostgreSQL y para incluir JSON, fechas y números.
    cursor.execute(f"SELECT {', '.join(qn(c) for c in columns)} FROM {qn(table_name)}")
    for row in cursor.fetchall():
     values=dict(zip(columns,row)); matched=[]
     for col,value in values.items():
      if value is None: continue
      if isinstance(value,(dict,list,tuple)):
       text=json.dumps(value,ensure_ascii=False,default=str)
      else:
       text=str(value)
      if needle in text.casefold(): matched.append({"field":col,"value":text})
     if matched:
      row_id=values.get("id") or values.get("pk") or values.get("internal_id") or values.get("key") or "—"
      # Resumen corto de la fila, evitando repetir contenidos enormes (JSON/logs).
      summary=[]
      for col,value in values.items():
       if value in (None,"") or col in {"id"}: continue
       text=json.dumps(value,ensure_ascii=False,default=str) if isinstance(value,(dict,list,tuple)) else str(value)
       if len(text)>140: text=text[:137]+"…"
       summary.append(f"{col}: {text}")
       if len(summary)>=4: break
      result={"table":table_name,"table_label":_global_table_label(table_name),"row_id":row_id,"matches":matched,"summary":" · ".join(summary)}
      if table_name==InventoryRecord._meta.db_table and values.get("id"):
       result["record_pk"]=values["id"]
      results.append(result)
   except Exception:
    continue
 return results
def client_ip(request):
 remote=(request.META.get("REMOTE_ADDR") or "").strip()
 if remote in {"127.0.0.1","::1"}:
  forwarded=(request.META.get("HTTP_X_FORWARDED_FOR") or "").split(",")[0].strip()
  if forwarded: return forwarded
 return remote or "0.0.0.0"
def safe_return(request):
 target=request.POST.get("next","")
 return target if url_has_allowed_host_and_scheme(target,{request.get_host()},request.is_secure()) else "dashboard"
def is_operator(request): return bool(request.user.is_authenticated and request.user.is_superuser)
def is_admin(user): return bool(user.is_authenticated and user.is_staff and not user.is_superuser)
def role_required(role):
 def decorator(view):
  @wraps(view)
  @login_required
  def wrapped(request,*args,**kwargs):
   allowed=is_operator(request) if role=="operator" else (is_operator(request) or is_admin(request.user))
   if not allowed:
    AuditLog.objects.create(user=request.user,action="access_denied",object_type="Route",object_id=request.path,details={"required":role,"ip":client_ip(request)})
    return HttpResponseForbidden("No tiene permiso para realizar esta operación.")
   return view(request,*args,**kwargs)
  return wrapped
 return decorator
admin_required=role_required("admin"); operator_required=role_required("operator")

@login_required
def certificate_help(request):
 cert_path=Path(settings.BASE_DIR)/"certs"/"PULSIA-Inventario-Root-CA.crt"
 cert_ready=cert_path.is_file() and cert_path.stat().st_size>0
 fingerprint=""
 if cert_ready:
  fingerprint=hashlib.sha256(cert_path.read_bytes()).hexdigest().upper()
  fingerprint=":".join(fingerprint[i:i+2] for i in range(0,len(fingerprint),2))
 return render(request,"inventory/certificate_help.html",{
  "cert_ready":cert_ready,
  "cert_fingerprint":fingerprint,
  "access_host":request.get_host(),
 })

@login_required
def certificate_download(request):
 cert_path=Path(settings.BASE_DIR)/"certs"/"PULSIA-Inventario-Root-CA.crt"
 if not cert_path.is_file():
  return HttpResponse("La CA PULSIA todavía no está disponible. Ejecute de nuevo el instalador HTTPS o revise el servicio Caddy.",status=503,content_type="text/plain; charset=utf-8")
 response=FileResponse(open(cert_path,"rb"),content_type="application/x-x509-ca-cert")
 response["Content-Disposition"]='attachment; filename="PULSIA-Inventario-Root-CA.crt"'
 response["Cache-Control"]="no-store"
 return response

def sensitive_field(field):
 return bool(field.is_destination_sn or field.is_technician or field.field_type=="date" or "fecha" in field.key or "fehca" in field.key)

def is_motherboard_table(table):
 text=f"{table.name} {table.slug}".casefold()
 return "placa" in text and "base" in text

def _record_field_value(record,predicate):
 for field in record.table.inventory_fields.all():
  if predicate(field):
   value=record.data.get(field.key)
   if value not in (None,""): return str(value).strip()
 return ""

def motherboard_unavailable(record):
 state=_record_field_value(record,lambda f:"estado" in f.key.casefold() or "estado" in f.name.casefold())
 if state.casefold()=="ko": return True
 technician=_record_field_value(record,lambda f:f.is_technician or "tecnic" in f.key.casefold() or "técnic" in f.name.casefold() or "tecnic" in f.name.casefold()) or str(record.current_technician or "").strip()
 date_value=_record_field_value(record,lambda f:f.field_type=="date" or "fecha" in f.key.casefold() or "fecha" in f.name.casefold() or "fehca" in f.key.casefold())
 sn_value=_record_field_value(record,lambda f:f.is_destination_sn or f.key.casefold()=="sn" or "serial" in f.key.casefold() or "serie" in f.name.casefold()) or str(record.current_sn or "").strip()
 return bool(technician and date_value and sn_value)

def completed_delivery(record):
 if is_motherboard_table(record.table): return motherboard_unavailable(record)
 fields=list(record.table.inventory_fields.all()); has_date=any(record.data.get(f.key) not in (None,"") for f in fields if f.field_type=="date" or "fecha" in f.key or "fehca" in f.key)
 return bool(str(record.current_sn or "").strip() and str(record.current_technician or "").strip() and has_date)

def record_is_available(record):
 if record.status=="scrapped": return False
 if is_motherboard_table(record.table): return not motherboard_unavailable(record)
 return record.status not in {"reserved","loaned","assigned"} and not completed_delivery(record)

def visible_to_normal(record): return record_is_available(record)

def public_record_text(record):
 fields=record.table.inventory_fields.all(); allowed={f.key for f in fields if not sensitive_field(f)}
 return json.dumps({k:v for k,v in record.data.items() if k in allowed},ensure_ascii=False,default=str)

def apply_delivery_data(record,technician,sn,occurred_at):
 data=record.data.copy()
 motherboard=is_motherboard_table(record.table)
 for field in record.table.inventory_fields.filter(is_destination_sn=True):
  if not motherboard: data[field.key]=sn
 for field in record.table.inventory_fields.filter(is_technician=True): data[field.key]=technician
 for field in record.table.inventory_fields.filter(is_primary=False):
  if field.field_type=="date" or "fecha" in field.key or "fehca" in field.key: data[field.key]=occurred_at.date().isoformat()
 record.data=data
 if not motherboard: record.current_sn=sn
 record.current_technician=technician
 record.status="assigned" if (not motherboard or motherboard_unavailable(record)) else "available"

def gestor_bootstrap_login(request, token):
 raw=(token or "").strip()
 if not raw:
  return HttpResponseForbidden("Acceso inicial no válido.")
 digest=hashlib.sha256(raw.encode("utf-8")).hexdigest()
 now=timezone.now()
 profile=UserProfile.objects.select_related("user").filter(bootstrap_token_hash=digest,bootstrap_expires_at__gte=now,bootstrap_used_at__isnull=True,user__is_active=True,user__is_superuser=True).first()
 if not profile:
  return HttpResponseForbidden("Este acceso inicial no es válido, ya fue utilizado o ha caducado.")
 user=profile.user
 if user.has_usable_password():
  return HttpResponseForbidden("El Gestor ya tiene una contraseña establecida.")
 profile.bootstrap_used_at=now; profile.bootstrap_token_hash=""; profile.bootstrap_expires_at=None; profile.save(update_fields=["bootstrap_used_at","bootstrap_token_hash","bootstrap_expires_at","updated_at"])
 login(request,user,backend="django.contrib.auth.backends.ModelBackend")
 AuditLog.objects.create(user=user,action="gestor_bootstrap_login",object_type="User",object_id=str(user.pk),details={"ip":client_ip(request)})
 messages.warning(request,"Acceso inicial del Gestor. Establezca ahora su contraseña desde esta pantalla.")
 return redirect("users_panel")

def auto_register_login(request):
 if request.user.is_authenticated: return redirect("dashboard")
 error=""; mode=request.POST.get("action","login")
 if request.method=="POST":
  username=request.POST.get("username","").strip().lower(); password=request.POST.get("password",""); User=get_user_model(); existing=User.objects.filter(username__iexact=username).first() if username else None
  if mode=="request_reset":
   if existing and existing.is_active:
    profile,_=UserProfile.objects.get_or_create(user=existing); profile.password_reset_requested_at=timezone.now(); profile.password_reset_authorized_at=None; profile.save(update_fields=["password_reset_requested_at","password_reset_authorized_at","updated_at"])
    messages.success(request,"Solicitud enviada al Gestor. Cuando la autorice podrás establecer una nueva contraseña desde esta pantalla.")
   else: messages.success(request,"Si la cuenta existe y está activa, la solicitud quedará disponible para el Gestor.")
   return redirect("login")

  if mode=="set_reset_password":
   confirm=request.POST.get("password_confirm","")
   if not existing: error="La cuenta no está disponible para restablecimiento."
   else:
    profile,_=UserProfile.objects.get_or_create(user=existing)
    if not profile.password_reset_authorized_at or existing.has_usable_password(): error="El Gestor todavía no ha autorizado el restablecimiento de esta cuenta."
    elif len(password)<4: error="La nueva contraseña debe tener al menos 4 caracteres."
    elif password!=confirm: error="Las contraseñas no coinciden."
    else:
     existing.set_password(password); existing.save(update_fields=["password"])
     profile.password_reset_authorized_at=None; profile.password_reset_requested_at=None; profile.must_change_password=False; profile.save()
     if not login_allowed(existing,request):
      error="Acceso fuera del horario autorizado. La cuenta ha quedado bloqueada hasta que el Gestor la revise."
     else:
      login(request,existing,backend="django.contrib.auth.backends.ModelBackend"); register_login_session(existing,request)
      AuditLog.objects.create(user=existing,action="password_reset_completed",object_type="User",object_id=str(existing.pk),details={"ip":client_ip(request)})
      messages.success(request,"Contraseña establecida correctamente."); return redirect("dashboard")
  else:
   if not username or len(password)<4: error="El usuario es obligatorio y la contraseña debe tener al menos 4 caracteres."
   elif existing:
    profile,_=UserProfile.objects.get_or_create(user=existing)
    if not existing.is_active: error="La cuenta está bloqueada o no disponible. El Gestor debe revisarla."
    elif profile.password_reset_authorized_at and not existing.has_usable_password(): error="El Gestor ha autorizado el restablecimiento. Usa el formulario 'Establecer nueva contraseña'."
    else:
     user=authenticate(request,username=existing.username,password=password)
     if not user:
      error="El usuario o la contraseña no son correctos."
     elif not user.is_active:
      error="La cuenta está bloqueada o no disponible. El Gestor debe revisarla."
     elif not login_allowed(user,request):
      error="Acceso fuera del horario autorizado. La cuenta ha quedado bloqueada indefinidamente hasta revisión del Gestor."
     else:
      login(request,user); register_login_session(user,request)
      AuditLog.objects.create(user=user,action="login",object_type="User",object_id=str(user.pk),details={"ip":client_ip(request)})
      target=request.POST.get("next",""); return redirect(target if url_has_allowed_host_and_scheme(target,{request.get_host()},request.is_secure()) else "dashboard")
   elif username in RESERVED_NAMES:
    error="Ese nombre de usuario está reservado."
   else:
    user=User.objects.create_user(username=username,password=password)
    UserProfile.objects.create(user=user,created_ip=client_ip(request),role="guest")
    if not login_allowed(user,request):
     error="La cuenta se ha creado, pero el acceso se intentó fuera del horario permitido y ha quedado bloqueada para revisión del Gestor."
    else:
     login(request,user); register_login_session(user,request)
     AuditLog.objects.create(user=user,action="self_registered_guest",object_type="User",object_id=str(user.pk),details={"ip":client_ip(request)})
     messages.success(request,"Cuenta creada como Invitado. Puede usar el chat y solicitar acceso como Usuario."); return redirect("dashboard")
 return render(request,"registration/login.html",{"error":error,"next":request.POST.get("next","") or request.GET.get("next",""),"reset_username":request.POST.get("username","")})

@login_required
def request_access_upgrade(request):
 profile,_=UserProfile.objects.get_or_create(user=request.user)
 if not profile.is_guest:
  messages.info(request,"Su cuenta ya dispone de acceso como Usuario o superior.")
  return redirect("dashboard")
 existing=AccessUpgradeRequest.objects.filter(user=request.user).first()
 if request.method=="POST":
  if existing:
   if existing.status=="pending": messages.info(request,"Su solicitud ya está pendiente de revisión por el Gestor.")
   elif existing.status=="approved": messages.info(request,"Su solicitud ya fue aprobada.")
   else: messages.error(request,"La solicitud fue denegada.")
  else:
   ip=client_ip(request)
   AccessUpgradeRequest.objects.create(user=request.user,requested_ip=ip)
   AuditLog.objects.create(user=request.user,action="guest_upgrade_requested",object_type="User",object_id=str(request.user.pk),details={"ip":ip})
   messages.success(request,"Solicitud enviada al Gestor.")
  return redirect("dashboard")
 return redirect("dashboard")

@login_required
def change_required_password(request):
 profile,_=UserProfile.objects.get_or_create(user=request.user); form=PasswordChangeRequiredForm(request.POST or None)
 if request.method=="POST" and form.is_valid():
  request.user.set_password(form.cleaned_data["password"]); request.user.save(); profile.must_change_password=False; profile.save(); login(request,request.user,backend="django.contrib.auth.backends.ModelBackend"); messages.success(request,"Contraseña actualizada."); return redirect("dashboard")
 return render(request,"inventory/form.html",{"form":form,"title":"Establecer nueva contraseña","submit":"Guardar contraseña"})

@login_required
@ensure_csrf_cookie
def dashboard(request):
 profile,_=UserProfile.objects.get_or_create(user=request.user)
 if profile.is_guest:
  access_request=AccessUpgradeRequest.objects.filter(user=request.user).first()
  return render(request,"inventory/guest_dashboard.html",{"access_request":access_request,"is_guest":True})
 # Un usuario ya aceptado entra directamente en su pizarra de producción.
 # Administradores y Gestor mantienen el resumen de inventario como pantalla de trabajo.
 if not request.user.is_staff and not request.user.is_superuser and request.GET.get("view") != "inventory":
  return redirect("production_board")
 q=request.GET.get("q","").strip(); global_q=request.GET.get("global_q","").strip(); restricted=not request.user.is_staff; qs=InventoryRecord.objects.select_related("table").prefetch_related("table__inventory_fields").order_by("-updated_at")
 if q:
  records=[]
  for record in qs:
   if restricted and not visible_to_normal(record): continue
   searchable=record.internal_id+" "+(public_record_text(record) if restricted else record.current_sn+" "+record.current_technician+" "+json.dumps(record.data,ensure_ascii=False,default=str))
   if q.casefold() in searchable.casefold(): records.append(record)
   if len(records)>=100: break
 else:
  records=[]
  for record in qs:
   if not restricted or visible_to_normal(record): records.append(record)
   if len(records)>=50: break
 tables=list(InventoryTable.objects.filter(active=True).prefetch_related("records","inventory_fields")); table_cards=[]
 for table in tables:
  count=sum(1 for record in table.records.all() if not restricted or visible_to_normal(record)); table_cards.append({"table":table,"count":count})
 visible_all=sum(card["count"] for card in table_cards); visible_available=sum(1 for record in qs if (not restricted or visible_to_normal(record)) and record_is_available(record))
 global_results=global_database_search(global_q) if global_q and request.user.is_staff else []
 context={"records":records,"tables":tables,"table_cards":table_cards,"restricted_user":restricted,"query":q,"global_query":global_q,"global_results":global_results,"is_admin":request.user.is_staff,"is_operator":request.user.is_superuser,"totals":{"all":visible_all,"available":visible_available,"assigned":InventoryRecord.objects.filter(status="assigned").count() if not restricted else 0,"incidents":Incident.objects.filter(status="pending").count() if not restricted else 0}}
 return render(request,"inventory/dashboard.html",context)

@admin_required
def add_item(request):
 table_slug=request.POST.get("table_slug") or request.GET.get("table")
 if not table_slug: return render(request,"inventory/select_table.html",{"tables":InventoryTable.objects.filter(active=True),"title":"Seleccionar tabla para el alta"})
 table=get_object_or_404(InventoryTable,slug=table_slug,active=True); form=DynamicRecordForm(table,request.POST or None)
 if request.method=="POST" and form.is_valid():
  try:
   with transaction.atomic():
    record=form.save(request.user)
    RecordMovement.objects.create(record=record,movement_type="entry",reason="Alta inicial",registered_by=request.user)
    AuditLog.objects.create(user=request.user,action="record_created",object_type=table.name,object_id=record.internal_id)
  except IntegrityError as exc:
   form.add_error(None,"No se pudo guardar porque el identificador generado ya existe. Vuelva a intentarlo.")
   AuditLog.objects.create(user=request.user,action="record_save_error",object_type=table.name,details={"error":str(exc)[:500]})
  except Exception as exc:
   # El usuario recibe un error de formulario en vez de una página HTTP 500.
   form.add_error(None,f"No se pudo guardar el objeto: {exc}")
   try: AuditLog.objects.create(user=request.user,action="record_save_error",object_type=table.name,details={"error":str(exc)[:500]})
   except Exception: pass
  else:
   print_job=print_identifier(record.internal_id,request.user,2)
   if print_job.status=="printed": messages.success(request,f"Objeto añadido. Identificador generado e impreso por duplicado: {record.internal_id}.")
   else: messages.warning(request,f"Objeto añadido con identificador {record.internal_id}, pero no se pudo imprimir: {print_job.error}")
   return redirect("table_view",slug=table.slug)
 return render(request,"inventory/record_form.html",{"form":form,"table":table,"title":f"Alta en {table.name}","submit":"Guardar objeto"})

@admin_required
def edit_item(request,pk):
 record=get_object_or_404(InventoryRecord.objects.select_related("table"),pk=pk); form=DynamicRecordForm(record.table,request.POST or None,instance=record)
 if request.method=="POST" and form.is_valid():
  try:
   with transaction.atomic():
    record=form.save(request.user)
    AuditLog.objects.create(user=request.user,action="record_modified",object_type=record.table.name,object_id=record.internal_id)
  except Exception as exc:
   form.add_error(None,f"No se pudieron guardar los cambios: {exc}")
   try: AuditLog.objects.create(user=request.user,action="record_save_error",object_type=record.table.name,object_id=record.internal_id,details={"error":str(exc)[:500]})
   except Exception: pass
  else:
   messages.success(request,"Objeto modificado."); return redirect("table_view",slug=record.table.slug)
 return render(request,"inventory/record_form.html",{"form":form,"table":record.table,"title":f"Modificar {record.internal_id}","submit":"Guardar cambios"})

@admin_required
def assign_item(request):
 form=RecordAssignmentForm(request.POST or None,initial={"occurred_at":timezone.localtime().strftime("%Y-%m-%dT%H:%M")})
 if request.method=="POST" and form.is_valid():
  d=form.cleaned_data; record=d["record"]; RecordMovement.objects.create(record=record,movement_type="assign",occurred_at=d["occurred_at"],technician_name=d["technician"],destination_sn=d["destination_sn"],reason=d["reason"],registered_by=request.user); apply_delivery_data(record,d["technician"],d["destination_sn"],d["occurred_at"]); data=record.data.copy()
  if d["reason"]:
   note=record.table.inventory_fields.filter(key__icontains="nota",is_primary=False).first()
   if note: data[note.key]=d["reason"]
  record.data=data; record.save(); messages.success(request,"Entrega registrada."); return redirect("trace")
 return render(request,"inventory/assignment_form.html",{"form":form,"title":"Entregar objeto a técnico","submit":"Registrar entrega"})

@admin_required
def object_search(request):
 q=request.GET.get("q","").strip()
 if len(q)<2: return JsonResponse({"results":[]})
 candidates=InventoryRecord.objects.exclude(status="scrapped").select_related("table").order_by("internal_id")
 records=[]
 q_fold=q.casefold()
 for record in candidates[:2000]:
  haystack=" ".join([record.internal_id,record.current_sn,record.current_technician,json.dumps(record.data,ensure_ascii=False,default=str)])
  if q_fold in haystack.casefold(): records.append(record)
  if len(records)>=20: break
 results=[]
 for record in records:
  details=[str(v) for v in record.data.values() if v not in (None,"")][:3]
  results.append({"pk":record.pk,"id":record.internal_id,"table":record.table.name,"description":" · ".join(details),"status":record.get_status_display(),"sn":record.current_sn,"technician":record.current_technician})
 return JsonResponse({"results":results})

@admin_required
def scrap_item(request):
 form=RecordScrapForm(request.POST or None,initial={"occurred_at":timezone.localtime().strftime("%Y-%m-%dT%H:%M")})
 if request.method=="POST" and form.is_valid():
  d=form.cleaned_data; record=d["record"]; RecordMovement.objects.create(record=record,movement_type="scrap",occurred_at=d["occurred_at"],technician_name=record.current_technician,destination_sn=record.current_sn,reason=d["reason"],registered_by=request.user); data=record.data.copy()
  state=record.table.inventory_fields.filter(key__icontains="estado",is_primary=False).first(); note=record.table.inventory_fields.filter(key__icontains="nota",is_primary=False).first()
  if state: data[state.key]="Baja / merma"
  if note: data[note.key]=d["reason"]
  record.data=data; record.status="scrapped"; record.save(); messages.success(request,"Baja registrada conservando la trazabilidad."); return redirect("dashboard")
 return render(request,"inventory/form.html",{"form":form,"title":"Dar de baja · Merma","submit":"Confirmar baja"})

@admin_required
def productivity_report(request):
 now=timezone.localtime()
 preset=request.GET.get("preset","today")
 start_text=request.GET.get("start","").strip(); end_text=request.GET.get("end","").strip(); user_id=request.GET.get("user","").strip(); granularity=request.GET.get("granularity","hour")
 if preset=="hour": start=now-timedelta(hours=1); end=now
 elif preset=="yesterday":
  day=now.date()-timedelta(days=1); start=timezone.make_aware(datetime.combine(day,datetime.min.time())); end=start+timedelta(days=1)
 elif preset=="week": start=now-timedelta(days=7); end=now
 elif preset=="month": start=now-timedelta(days=30); end=now
 elif preset=="custom":
  try: start=timezone.make_aware(datetime.fromisoformat(start_text)) if start_text else now.replace(hour=0,minute=0,second=0,microsecond=0)
  except ValueError: start=now.replace(hour=0,minute=0,second=0,microsecond=0)
  try: end=timezone.make_aware(datetime.fromisoformat(end_text)) if end_text else now
  except ValueError: end=now
 else: start=now.replace(hour=0,minute=0,second=0,microsecond=0); end=now; preset="today"
 qs=InventoryRecord.objects.filter(created_at__gte=start,created_at__lte=end).select_related("created_by","table")
 if user_id.isdigit(): qs=qs.filter(created_by_id=int(user_id))
 by_user=list(qs.values("created_by_id","created_by__username").annotate(total=Count("id")).order_by("-total","created_by__username"))
 trunc={"hour":TruncHour,"day":TruncDay,"week":TruncWeek,"month":TruncMonth}.get(granularity,TruncHour)
 timeline=list(qs.annotate(period=trunc("created_at")).values("period").annotate(total=Count("id")).order_by("period"))
 records=qs.order_by("-created_at")[:500]
 users=get_user_model().objects.filter(inventory_records_created__isnull=False).distinct().order_by("username")
 return render(request,"inventory/productivity.html",{"total_all":InventoryRecord.objects.count(),"total_period":qs.count(),"by_user":by_user,"timeline":timeline,"records":records,"users":users,"selected_user":user_id,"preset":preset,"granularity":granularity,"start_text":start_text,"end_text":end_text,"period_start":start,"period_end":end})

def _client_row_snapshot(row):
 return {"internal_id":row.internal_id,"brand":row.brand,"model_reference":row.model_reference,"component":row.component,"reference":row.reference,"units_pending":row.units_pending,"units_stock":row.units_stock,"units_sent":row.units_sent,"unit_price":str(row.unit_price),"total_price":str(row.total_price),"client":row.client,"observations":row.observations,"extra_data":row.extra_data.copy()}

def _recalculate_client_row(row):
 from decimal import Decimal,ROUND_HALF_UP
 row.total_price=(Decimal(row.units_pending+row.units_stock+row.units_sent)*row.unit_price).quantize(Decimal("0.01"),rounding=ROUND_HALF_UP)

@admin_required
def client_batches(request,sheet_id=None):
 sheets=ClientBatchSheet.objects.filter(active=True).order_by("client","concept","position","name")
 sheet=get_object_or_404(ClientBatchSheet,pk=sheet_id,active=True) if sheet_id else sheets.first()
 allowed_tabs={"records","add","sheets","fields","history"}
 active_tab=request.GET.get("tab","records")
 if active_tab not in allowed_tabs: active_tab="records"
 sheet_form=ClientBatchSheetForm(prefix="sheet")
 current_sheet_form=ClientBatchSheetForm(instance=sheet,prefix="sheet_edit") if sheet else None
 field_form=ClientBatchFieldForm(prefix="field")
 row_form=ClientBatchRowForm(prefix="row",sheet=sheet) if sheet else None
 def _batch_redirect(tab,selected=None):
  target=selected or sheet
  if target:
   base=reverse("client_batches_sheet",kwargs={"sheet_id":target.pk})
   return redirect(base if tab=="records" else f"{base}?tab={tab}")
  base=reverse("client_batches")
  return redirect(base if tab=="records" else f"{base}?tab={tab}")
 if request.method=="POST":
  action=request.POST.get("action","")
  if action=="create_sheet":
   sheet_form=ClientBatchSheetForm(request.POST,prefix="sheet")
   if sheet_form.is_valid():
    obj=sheet_form.save(commit=False); obj.created_by=request.user; obj.save(); ClientBatchChange.objects.create(sheet=obj,action="sheet_created",after={"name":obj.name,"client":obj.client,"concept":obj.concept},changed_by=request.user); messages.success(request,"Hoja interna creada."); return _batch_redirect("records",obj)
   active_tab="sheets"
  elif sheet and action=="update_sheet":
   before={"name":sheet.name,"client":sheet.client,"concept":sheet.concept}; current_sheet_form=ClientBatchSheetForm(request.POST,instance=sheet,prefix="sheet_edit")
   if current_sheet_form.is_valid():
    obj=current_sheet_form.save(); ClientBatchChange.objects.create(sheet=obj,action="sheet_modified",before=before,after={"name":obj.name,"client":obj.client,"concept":obj.concept},changed_by=request.user); messages.success(request,"Datos de la hoja actualizados."); return _batch_redirect("sheets",obj)
   active_tab="sheets"
  elif sheet and action=="create_field":
   field_form=ClientBatchFieldForm(request.POST,prefix="field")
   if field_form.is_valid():
    field=field_form.save(commit=False); field.sheet=sheet; field.created_by=request.user; base=re.sub(r"[^a-z0-9_]+","_",field.name.lower()).strip("_") or "campo"; key=base; n=2
    while ClientBatchField.objects.filter(sheet=sheet,key=key).exists(): key=f"{base}_{n}"; n+=1
    field.key=key; field.position=sheet.custom_fields.count()+1; field.save(); ClientBatchChange.objects.create(sheet=sheet,field=field,action="field_created",after={"name":field.name,"key":field.key,"type":field.field_type},changed_by=request.user); messages.success(request,"Campo añadido."); return _batch_redirect("fields")
   active_tab="fields"
  elif sheet and action=="update_field":
   field=get_object_or_404(ClientBatchField,pk=request.POST.get("field_id"),sheet=sheet); before={"name":field.name,"type":field.field_type,"active":field.active}; form=ClientBatchFieldForm(request.POST,instance=field,prefix="field_edit")
   if form.is_valid():
    field=form.save(); ClientBatchChange.objects.create(sheet=sheet,field=field,action="field_modified",before=before,after={"name":field.name,"type":field.field_type,"active":field.active},changed_by=request.user); messages.success(request,"Campo modificado."); return _batch_redirect("fields")
   active_tab="fields"
  elif sheet and action=="disable_field":
   field=get_object_or_404(ClientBatchField,pk=request.POST.get("field_id"),sheet=sheet); before={"name":field.name,"active":field.active}; field.active=False; field.save(update_fields=["active","updated_at"]); ClientBatchChange.objects.create(sheet=sheet,field=field,action="field_disabled",before=before,after={"name":field.name,"active":False},changed_by=request.user); messages.success(request,"Campo retirado de la hoja; el histórico se conserva."); return _batch_redirect("fields")
  elif sheet and action=="create_row":
   row_form=ClientBatchRowForm(request.POST,prefix="row",sheet=sheet)
   if row_form.is_valid():
    row=row_form.save(commit=False); row.created_by=request.user; row.updated_by=request.user; _recalculate_client_row(row)
    try:
     with transaction.atomic():
      locked_sheet=ClientBatchSheet.objects.select_for_update().get(pk=sheet.pk)
      next_id=locked_sheet.next_row_number
      while ClientBatchRow.objects.filter(sheet=locked_sheet,internal_id=str(next_id)).exists(): next_id+=1
      row.sheet=locked_sheet; row.internal_id=str(next_id); locked_sheet.next_row_number=next_id+1; locked_sheet.save(update_fields=["next_row_number","updated_at"]); row.save()
      ClientBatchChange.objects.create(sheet=sheet,row=row,action="row_created",after=_client_row_snapshot(row),changed_by=request.user); messages.success(request,"Registro añadido a la hoja."); return _batch_redirect("records")
    except IntegrityError: row_form.add_error(None,"No se pudo asignar un ID incremental único. Inténtelo de nuevo.")
   active_tab="add"
  elif sheet and action=="update_cell":
   row=get_object_or_404(ClientBatchRow,pk=request.POST.get("row_id"),sheet=sheet); field=request.POST.get("field",""); value=request.POST.get("value","").strip(); before=_client_row_snapshot(row)
   try:
    with transaction.atomic():
     row=ClientBatchRow.objects.select_for_update().get(pk=row.pk); before=_client_row_snapshot(row)
     if field in {"units_pending","units_stock","units_sent"}:
      new=int(value or 0)
      if new<0: raise ValueError("Las unidades no pueden ser negativas.")
      if field=="units_stock":
       delta=new-row.units_stock
       if row.units_pending-delta<0: raise ValueError("No hay suficientes UD pendientes para aumentar el stock.")
       row.units_stock=new; row.units_pending-=delta
      elif field=="units_sent":
       delta=new-row.units_sent
       if row.units_stock-delta<0: raise ValueError("No hay suficiente stock para aumentar las UD enviadas.")
       row.units_sent=new; row.units_stock-=delta
      else: row.units_pending=new
     elif field=="unit_price":
      from decimal import Decimal,InvalidOperation
      try: new=Decimal(value.replace(",","."))
      except InvalidOperation: raise ValueError("Precio unitario no válido.")
      if new<0: raise ValueError("El precio no puede ser negativo.")
      row.unit_price=new
     elif field in {"brand","model_reference","component","reference","client","observations"}:
      setattr(row,field,value)
     elif field.startswith("extra:"):
      key=field.split(":",1)[1]; definition=get_object_or_404(ClientBatchField,sheet=sheet,key=key,active=True); data=row.extra_data.copy()
      if definition.field_type=="number" and value:
       try: value=float(value.replace(",","."))
       except ValueError: raise ValueError("Valor numérico no válido.")
      elif definition.field_type=="bool": value=value.lower() in {"1","true","sí","si","yes"}
      data[key]=value; row.extra_data=data
     else: raise ValueError("Campo no editable.")
     row.updated_by=request.user; _recalculate_client_row(row); row.save(); ClientBatchChange.objects.create(sheet=sheet,row=row,action="row_modified",before=before,after=_client_row_snapshot(row),changed_by=request.user)
    messages.success(request,"Cambio guardado.")
   except (ValueError,IntegrityError) as exc: messages.error(request,str(exc) or "No se pudo guardar el cambio.")
   return _batch_redirect("records")
 rows=sheet.rows.select_related("created_by","updated_by").order_by("internal_id") if sheet else ClientBatchRow.objects.none()
 fields=sheet.custom_fields.filter(active=True) if sheet else ClientBatchField.objects.none()
 changes=sheet.changes.select_related("changed_by","row","field")[:150] if sheet else ClientBatchChange.objects.none()
 if sheet:
  row_form=row_form or ClientBatchRowForm(prefix="row",sheet=sheet)
 return render(request,"inventory/client_batches.html",{"sheets":sheets,"sheet":sheet,"fields":fields,"rows":rows,"changes":changes,"sheet_form":sheet_form,"current_sheet_form":current_sheet_form,"field_form":field_form,"row_form":row_form,"active_tab":active_tab})

@admin_required
def trace(request):
 q=request.GET.get("q","").strip(); mode=request.GET.get("mode","sn"); movements=RecordMovement.objects.select_related("record","record__table","registered_by").none()
 if q:
  movements=RecordMovement.objects.select_related("record","record__table","registered_by").filter(technician_name__icontains=q) if mode=="technician" else RecordMovement.objects.select_related("record","record__table","registered_by").filter(destination_sn__icontains=q)
 return render(request,"inventory/trace.html",{"query":q,"mode":mode,"movements":movements.order_by("-occurred_at")})

@admin_required
def entry_report(request):
 selected_text=request.GET.get("date","").strip(); selected_date=None; records=[]; by_user=[]; error=""
 if selected_text:
  try:
   selected_date=date.fromisoformat(selected_text)
   records=list(InventoryRecord.objects.filter(created_at__date=selected_date).select_related("table","created_by").order_by("created_at","table__name","internal_id"))
   totals={}
   for record in records:
    username=record.created_by.get_username(); totals[username]=totals.get(username,0)+1
   by_user=sorted(totals.items(),key=lambda item:(-item[1],item[0].casefold()))
  except ValueError: error="Introduzca una fecha válida."
 return render(request,"inventory/entry_report.html",{"selected_text":selected_text,"selected_date":selected_date,"records":records,"by_user":by_user,"total":len(records),"error":error})

@login_required
def record_detail(request,pk):
 record=get_object_or_404(InventoryRecord.objects.select_related("table").prefetch_related("table__inventory_fields"),pk=pk); restricted=not request.user.is_staff
 if restricted and not visible_to_normal(record): return HttpResponseForbidden("Este objeto ya no está disponible.")
 fields=[]
 for field in record.table.inventory_fields.all():
  if restricted and sensitive_field(field): continue
  fields.append({"name":field.name,"value":record.internal_id if field.is_primary else record.data.get(field.key,"")})
 return JsonResponse({"id":record.internal_id,"table":record.table.name,"status":record.get_status_display(),"fields":fields,"can_reserve":record_is_available(record) and not Reservation.objects.filter(record=record,status__in=["pending","accepted"]).exists(),"reserve_url":reverse("reserve_record",kwargs={"pk":record.pk})})

@login_required
def reserve_record(request,pk):
 if request.method!="POST": return HttpResponse(status=405)
 form=ReservationForm(request.POST)
 if not form.is_valid(): messages.error(request,"Indique un destino y un SN de destino válidos."); return redirect(safe_return(request))
 try:
  with transaction.atomic():
   record=get_object_or_404(InventoryRecord.objects.select_for_update().select_related("table"),pk=pk)
   if not record_is_available(record) or Reservation.objects.filter(record=record,status__in=["pending","accepted"]).exists(): raise ValueError("El objeto ya no está disponible para solicitar.")
   reservation=Reservation.objects.create(record=record,requested_by=request.user,destination=form.cleaned_data["destination"],destination_sn=form.cleaned_data["destination_sn"],status="pending")
   RecordMovement.objects.create(record=record,movement_type="reserve",occurred_at=reservation.requested_at,technician_name=request.user.get_username(),destination_sn=reservation.destination_sn,reason=f"Solicitud de reserva para {reservation.get_destination_display()}",registered_by=request.user)
   AuditLog.objects.create(user=request.user,action="record_reserved",object_type=record.table.name,object_id=record.internal_id,details={"destination":reservation.destination,"sn":reservation.destination_sn})
 except (ValueError,IntegrityError) as exc: messages.error(request,str(exc) or "Otro usuario acaba de reservar este objeto.")
 else: messages.success(request,"Solicitud registrada. El objeto no cambia de estado hasta que se confirme la entrega.")
 return redirect(safe_return(request))

@login_required
def reservations_center(request):
 tab=request.GET.get("tab","search")
 if tab not in {"search","pending"}: tab="search"

 if request.method=="POST":
  action=request.POST.get("action","")
  reservation_id=request.POST.get("reservation_id")
  if action in {"accept","approve_deliver","reject"}:
   if not request.user.is_staff: return HttpResponseForbidden("Solo Gestor y Administradores pueden aprobar o rechazar solicitudes.")
   with transaction.atomic():
    reservation=get_object_or_404(Reservation.objects.select_for_update().select_related("record","record__table","requested_by"),pk=reservation_id,status="pending")
    record=InventoryRecord.objects.select_for_update().get(pk=reservation.record_id); now=timezone.now()
    if action=="reject":
     reservation.status="rejected"; reservation.resolved_by=request.user; reservation.resolved_at=now
     reservation.save(update_fields=["status","resolved_by","resolved_at"])
     RecordMovement.objects.create(record=record,movement_type="correction",occurred_at=now,technician_name=reservation.requested_by.get_username(),destination_sn=reservation.destination_sn,reason=f"Solicitud rechazada para {reservation.get_destination_display()}",registered_by=request.user)
     AuditLog.objects.create(user=request.user,action="reservation_rejected",object_type=record.table.name,object_id=record.internal_id,details={"reservation":reservation.pk})
     messages.success(request,"Solicitud rechazada. El objeto no ha cambiado de estado.")
    else:
     if not record_is_available(record): raise ValueError("El objeto ya no está disponible.")
     reservation.status="accepted"; reservation.accepted_by=request.user; reservation.accepted_at=now
     reservation.save(update_fields=["status","accepted_by","accepted_at"])
     AuditLog.objects.create(user=request.user,action="reservation_accepted",object_type=record.table.name,object_id=record.internal_id,details={"reservation":reservation.pk,"requested_by":reservation.requested_by.get_username()})
     if action=="approve_deliver":
      apply_delivery_data(record,reservation.requested_by.get_username(),reservation.destination_sn,now); record.save()
      reservation.status="delivered"; reservation.resolved_by=request.user; reservation.resolved_at=now
      reservation.save(update_fields=["status","resolved_by","resolved_at"])
      RecordMovement.objects.create(record=record,movement_type="assign",occurred_at=now,technician_name=reservation.requested_by.get_username(),destination_sn=reservation.destination_sn,reason=f"Aprobado y entregado para {reservation.get_destination_display()}",registered_by=request.user)
      AuditLog.objects.create(user=request.user,action="reservation_approved_and_delivered",object_type=record.table.name,object_id=record.internal_id,details={"reservation":reservation.pk,"confirmed_by":"staff"})
      messages.success(request,"Solicitud aprobada y objeto marcado como entregado.")
     else:
      messages.success(request,"Solicitud aprobada. El objeto mantiene su estado hasta confirmar la entrega.")
   return redirect(reverse("reservations_center")+"?tab=pending")

  if action in {"deliver","cancel","scrap"}:
   with transaction.atomic():
    reservation=get_object_or_404(Reservation.objects.select_for_update().select_related("record","record__table","requested_by"),pk=reservation_id,status="accepted")
    if not request.user.is_staff and reservation.requested_by_id!=request.user.id:
     return HttpResponseForbidden("Solo el receptor, Gestor o Administrador pueden confirmar la entrega.")
    record=InventoryRecord.objects.select_for_update().get(pk=reservation.record_id); now=timezone.now()
    if action=="deliver":
     if not record_is_available(record): raise ValueError("El objeto ya no está disponible para completar la entrega.")
     apply_delivery_data(record,reservation.requested_by.get_username(),reservation.destination_sn,now)
     reservation.status="delivered"; movement="assign"; reason=f"Objeto entregado para {reservation.get_destination_display()}"
     confirmation="staff" if request.user.is_staff else "recipient"
    elif action=="cancel":
     if not request.user.is_staff: return HttpResponseForbidden("Solo Gestor y Administradores pueden cancelar una aprobación.")
     reservation.status="cancelled"; movement="correction"; reason=f"Reserva para {reservation.get_destination_display()} cancelada"; confirmation="staff"
    else:
     if not request.user.is_staff: return HttpResponseForbidden("Solo Gestor y Administradores pueden registrar una merma.")
     reservation.status="scrapped"; record.status="scrapped"; data=record.data.copy()
     state=record.table.inventory_fields.filter(key__icontains="estado",is_primary=False).first(); note=record.table.inventory_fields.filter(key__icontains="nota",is_primary=False).first()
     if state: data[state.key]="KO"
     if note: data[note.key]=f"Merma desde reserva para {reservation.get_destination_display()}"
     record.data=data; movement="scrap"; reason=f"Merma desde reserva para {reservation.get_destination_display()}"; confirmation="staff"
    reservation.resolved_by=request.user; reservation.resolved_at=now
    reservation.save(update_fields=["status","resolved_by","resolved_at"]); record.save()
    RecordMovement.objects.create(record=record,movement_type=movement,occurred_at=now,technician_name=reservation.requested_by.get_username(),destination_sn=reservation.destination_sn,reason=reason,registered_by=request.user)
    AuditLog.objects.create(user=request.user,action=f"reservation_{reservation.status}",object_type=record.table.name,object_id=record.internal_id,details={"reservation":reservation.pk,"confirmation":confirmation})
   messages.success(request,"Reserva actualizada correctamente.")
   return redirect(reverse("reservations_center")+"?tab=pending")

  return HttpResponseForbidden("Acción no válida.")

 qs=Reservation.objects.select_related("record","record__table","requested_by","accepted_by","resolved_by").prefetch_related("record__table__inventory_fields").order_by("-requested_at")
 if not request.user.is_staff: qs=qs.filter(requested_by=request.user)

 pending_user=list(qs.filter(status="pending")[:500])
 approved_user=list(qs.filter(status="accepted")[:500])

 if tab=="pending":
  if request.user.is_staff:
   reservations=list(qs.filter(status__in=["pending","accepted"])[:500])
   pending_qs=Reservation.objects.filter(status="pending")
   already=set(ReservationView.objects.filter(user=request.user,reservation__in=pending_qs).values_list("reservation_id",flat=True))
   ReservationView.objects.bulk_create([ReservationView(user=request.user,reservation=item) for item in reservations if item.status=="pending" and item.pk not in already],ignore_conflicts=True)
  else:
   reservations=pending_user+approved_user
  return render(request,"inventory/reservations.html",{
   "tab":tab,"reservations":reservations,"pending_requests":pending_user,"approved_requests":approved_user,
   "pending_count":Reservation.objects.filter(status="pending").count() if request.user.is_staff else len(pending_user),
   "can_manage":request.user.is_staff,
  })

 status=request.GET.get("status","").strip(); technician=request.GET.get("technician","").strip(); sn=request.GET.get("sn","").strip(); text=request.GET.get("q","").strip(); table_id=request.GET.get("table","").strip(); date_from=request.GET.get("date_from","").strip(); date_to=request.GET.get("date_to","").strip(); delivered_only=request.GET.get("delivered_only")=="1"
 if delivered_only: qs=qs.filter(status="delivered")
 elif status: qs=qs.filter(status=status)
 if technician: qs=qs.filter(requested_by__username__icontains=technician)
 if sn: qs=qs.filter(destination_sn__icontains=sn)
 if table_id.isdigit(): qs=qs.filter(record__table_id=int(table_id))
 if date_from: qs=qs.filter(requested_at__date__gte=date_from)
 if date_to: qs=qs.filter(requested_at__date__lte=date_to)
 reservations=list(qs[:2000])
 field_filters=[]
 for i in range(1,4):
  fid=request.GET.get(f"field_{i}","").strip(); value=request.GET.get(f"value_{i}","").strip()
  if fid.isdigit() and value: field_filters.append((int(fid),value))
 if text or field_filters:
  field_map={f.pk:f for f in InventoryField.objects.select_related("table").filter(pk__in=[fid for fid,_ in field_filters])}
  filtered=[]
  for r in reservations:
   haystack=" ".join([r.record.internal_id,r.record.table.name,r.requested_by.get_username(),r.destination_sn,r.get_destination_display(),json.dumps(r.record.data,ensure_ascii=False,default=str)])
   if text and text.casefold() not in haystack.casefold(): continue
   ok=True
   for fid,value in field_filters:
    f=field_map.get(fid)
    if not f or f.table_id!=r.record.table_id or value.casefold() not in str(r.record.data.get(f.key,"")).casefold(): ok=False; break
   if ok: filtered.append(r)
  reservations=filtered
 return render(request,"inventory/reservations.html",{
  "tab":tab,"reservations":reservations,"pending_requests":pending_user,"approved_requests":approved_user,
  "pending_count":Reservation.objects.filter(status="pending").count() if request.user.is_staff else len(pending_user),
  "can_manage":request.user.is_staff,"tables":InventoryTable.objects.filter(active=True),
  "search_fields":InventoryField.objects.filter(is_primary=False,searchable=True).select_related("table").order_by("table__name","position"),
  "status_choices":Reservation.STATUS,"filters":request.GET,
 })

@login_required
def loans_center(request):
 tab=request.GET.get("tab","search")
 if tab not in {"search","pending","items"}: tab="search"
 if tab in {"pending","items"} and not request.user.is_staff: tab="search"
 if request.method=="POST":
  action=request.POST.get("action")
  if action=="request":
   form=LoanRequestForm(request.POST)
   if form.is_valid():
    try:
     with transaction.atomic():
      item=LoanItem.objects.select_for_update().get(pk=form.cleaned_data["item"].pk)
      if item.status!="available" or LoanRequest.objects.filter(item=item,status="pending").exists(): raise ValueError("El item ya no está disponible para solicitar.")
      req=LoanRequest.objects.create(item=item,requested_by=request.user,notes=form.cleaned_data["notes"]); item.status="pending"; item.save(update_fields=["status","updated_at"])
      AuditLog.objects.create(user=request.user,action="loan_requested",object_type="LoanItem",object_id=item.internal_id,details={"loan_request":req.pk})
     messages.success(request,"Solicitud de préstamo enviada. Queda pendiente de aceptación."); return redirect(reverse("loans_center")+"?tab=search")
    except (ValueError,IntegrityError) as exc: messages.error(request,str(exc))
  elif action in {"accept_request","reject_request"}:
   if not request.user.is_staff: return HttpResponseForbidden("Solo Gestor y Administradores pueden resolver solicitudes de préstamo.")
   with transaction.atomic():
    req=get_object_or_404(LoanRequest.objects.select_for_update().select_related("item","requested_by"),pk=request.POST.get("request_id"),status="pending")
    item=LoanItem.objects.select_for_update().get(pk=req.item_id); now=timezone.now()
    if action=="accept_request":
     if item.status not in {"pending","available"} or Loan.objects.filter(loan_item=item,returned_at__isnull=True).exists(): raise ValueError("El item ya no está disponible para préstamo.")
     req.status="accepted"; req.resolved_by=request.user; req.resolved_at=now; req.save(update_fields=["status","resolved_by","resolved_at"])
     loan=Loan.objects.create(loan_item=item,request=req,borrower=req.requested_by,technician_name=req.requested_by.get_username(),withdrawn_at=now,created_by=request.user,notes=req.notes); item.status="loaned"; item.save(update_fields=["status","updated_at"])
     AuditLog.objects.create(user=request.user,action="loan_request_accepted",object_type="LoanItem",object_id=item.internal_id,details={"loan_request":req.pk,"loan":loan.pk,"borrower":req.requested_by.get_username()})
     messages.success(request,"Préstamo aceptado. El item queda registrado como prestado.")
    else:
     req.status="rejected"; req.resolved_by=request.user; req.resolved_at=now; req.save(update_fields=["status","resolved_by","resolved_at"]); item.status="available"; item.save(update_fields=["status","updated_at"])
     AuditLog.objects.create(user=request.user,action="loan_request_rejected",object_type="LoanItem",object_id=item.internal_id,details={"loan_request":req.pk})
     messages.success(request,"Solicitud de préstamo rechazada. El item vuelve a estar disponible.")
   return redirect(reverse("loans_center")+"?tab=pending")
  elif action=="return":
   if not request.user.is_staff: return HttpResponseForbidden("Solo Gestor y Administradores pueden registrar devoluciones.")
   with transaction.atomic():
    loan=get_object_or_404(Loan.objects.select_for_update().select_related("record","record__table","loan_item","borrower"),pk=request.POST.get("loan_id"),returned_at__isnull=True); now=timezone.now(); loan.returned_at=now; loan.returned_by=request.user; loan.save(update_fields=["returned_at","returned_by"])
    if loan.loan_item_id:
     item=LoanItem.objects.select_for_update().get(pk=loan.loan_item_id); item.status="available"; item.save(update_fields=["status","updated_at"]); object_type="LoanItem"; object_id=item.internal_id
    else:
     record=loan.record; record.status="available"; record.save(update_fields=["status","updated_at"]); RecordMovement.objects.create(record=record,movement_type="loan_return",occurred_at=now,technician_name=loan.technician_name,reason=f"Devolución préstamo de {loan.borrower.get_username()}",registered_by=request.user); object_type=record.table.name; object_id=record.internal_id
    AuditLog.objects.create(user=request.user,action="loan_returned",object_type=object_type,object_id=object_id,details={"loan":loan.pk})
   messages.success(request,"Objeto devuelto. Vuelve a estar disponible."); return redirect(reverse("loans_center")+"?tab=search")
  elif action=="create_item":
   if not request.user.is_staff: return HttpResponseForbidden("Solo Gestor y Administradores pueden dar de alta items prestables.")
   form=LoanItemForm(request.POST)
   if form.is_valid():
    with transaction.atomic():
     existing=list(LoanItem.objects.select_for_update().values_list("internal_id",flat=True))
     maximum=0
     for value in existing:
      match=re.search(r"(\d+)$",value or "")
      if match: maximum=max(maximum,int(match.group(1)))
     item=form.save(commit=False); item.internal_id=f"PREST-{maximum+1:06d}"; item.created_by=request.user; item.status="available"; item.save()
    AuditLog.objects.create(user=request.user,action="loan_item_created",object_type="LoanItem",object_id=item.internal_id,details={"name":item.name}); messages.success(request,f"Item prestable dado de alta con identificador automático {item.internal_id}."); return redirect(reverse("loans_center")+"?tab=items")
  elif action=="item_status":
   if not request.user.is_staff: return HttpResponseForbidden("Solo Gestor y Administradores pueden gestionar items prestables.")
   item=get_object_or_404(LoanItem,pk=request.POST.get("item_id")); new_status=request.POST.get("status")
   if new_status not in {"available","out","retired"}: return HttpResponseForbidden("Estado no válido.")
   if item.status in {"loaned","pending"}: messages.error(request,"No se puede cambiar manualmente el estado mientras el item tenga un préstamo o solicitud pendiente.")
   else: item.status=new_status; item.save(update_fields=["status","updated_at"]); AuditLog.objects.create(user=request.user,action="loan_item_status",object_type="LoanItem",object_id=item.internal_id,details={"status":new_status}); messages.success(request,"Estado del item actualizado.")
   return redirect(reverse("loans_center")+"?tab=items")
 request_form=LoanRequestForm()
 item_form=LoanItemForm() if request.user.is_staff else None
 if tab=="pending":
  requests=LoanRequest.objects.filter(status="pending").select_related("item","requested_by").order_by("requested_at")
  return render(request,"inventory/loans.html",{"tab":tab,"loan_requests":requests,"can_manage":True,"pending_loan_requests_count":requests.count()})
 if tab=="items":
  items=LoanItem.objects.select_related("created_by").order_by("category","name","internal_id")
  return render(request,"inventory/loans.html",{"tab":tab,"loan_items":items,"item_form":item_form,"can_manage":True})
 qs=Loan.objects.select_related("record","record__table","loan_item","borrower","created_by","returned_by")
 if not request.user.is_staff: qs=qs.filter(borrower=request.user)
 status=request.GET.get("status","all"); q=request.GET.get("q","").strip(); date_from=request.GET.get("date_from","").strip(); date_to=request.GET.get("date_to","").strip()
 if status=="active": qs=qs.filter(returned_at__isnull=True)
 elif status=="returned": qs=qs.filter(returned_at__isnull=False)
 if date_from: qs=qs.filter(withdrawn_at__date__gte=date_from)
 if date_to: qs=qs.filter(withdrawn_at__date__lte=date_to)
 if q: qs=qs.filter(Q(loan_item__internal_id__icontains=q)|Q(loan_item__name__icontains=q)|Q(loan_item__serial_number__icontains=q)|Q(record__internal_id__icontains=q)|Q(borrower__username__icontains=q)|Q(technician_name__icontains=q)|Q(notes__icontains=q))
 available_items=LoanItem.objects.filter(status="available").order_by("category","name","internal_id")
 if q:
  available_items=available_items.filter(Q(internal_id__icontains=q)|Q(name__icontains=q)|Q(category__icontains=q)|Q(brand__icontains=q)|Q(model_reference__icontains=q)|Q(serial_number__icontains=q))
 own_pending=LoanRequest.objects.filter(requested_by=request.user,status="pending").select_related("item")
 return render(request,"inventory/loans.html",{"tab":tab,"loans":qs[:1000],"request_form":request_form,"available_items":available_items[:200],"own_pending_requests":own_pending,"can_manage":request.user.is_staff,"status_filter":status,"query":q,"date_from":date_from,"date_to":date_to})

@login_required
def chat_center(request,user_id=None):
 User=get_user_model(); users=User.objects.filter(is_active=True).exclude(pk=request.user.pk).order_by("username"); selected=None; conversation=ChatMessage.objects.none()
 if user_id:
  selected=get_object_or_404(users,pk=user_id)
  if request.method=="POST":
   body=request.POST.get("body","").strip()
   if not body: messages.error(request,"El mensaje no puede estar vacío.")
   elif len(body)>4000: messages.error(request,"El mensaje es demasiado largo.")
   else: ChatMessage.objects.create(sender=request.user,recipient=selected,body=body); return redirect("chat_conversation",user_id=selected.pk)
  ChatMessage.objects.filter(sender=selected,recipient=request.user,read_at__isnull=True).update(read_at=timezone.now())
  conversation=ChatMessage.objects.filter(Q(sender=request.user,recipient=selected)|Q(sender=selected,recipient=request.user)).select_related("sender","recipient")
 for account in users: account.unread_count=ChatMessage.objects.filter(sender=account,recipient=request.user,read_at__isnull=True).count()
 return render(request,"inventory/chat.html",{"chat_users":users,"selected_user":selected,"conversation":conversation})

@login_required
def notification_status(request):
 unread_messages=ChatMessage.objects.filter(recipient=request.user,read_at__isnull=True).count(); pending_reservations=0; pending_loan_requests=0
 profile,_=UserProfile.objects.get_or_create(user=request.user)
 if profile.is_guest:
  return JsonResponse({"unread_messages":unread_messages,"pending_reservations":0,"unseen_reservations":0,"pending_loan_requests":0,"active_loans":0,"pending_incidents":0,"security_red":0,"security_yellow":0})
 pending_incidents=0
 if request.user.is_staff:
  pending_reservations=Reservation.objects.filter(status="pending").count()
  pending_loan_requests=LoanRequest.objects.filter(status="pending").count()
  pending_incidents=Incident.objects.filter(status="pending").count()
 active_loans=Loan.objects.filter(returned_at__isnull=True).count() if request.user.is_staff else Loan.objects.filter(borrower=request.user,returned_at__isnull=True).count()
 security_red=SecurityAccessEvent.objects.filter(reviewed=False,level="RED").count() if request.user.is_staff else 0
 security_yellow=SecurityAccessEvent.objects.filter(reviewed=False,level="YELLOW").count() if request.user.is_staff else 0
 backup_disk_alert=0
 if request.user.is_superuser:
  try:
   cfg=BackupDiskConfig.objects.first()
   if cfg and (cfg.uuid or (cfg.mode=="local" and cfg.local_path)):
    st=request_storage_admin({"action":"status_backup_mount","uuid":cfg.uuid},timeout=1)
    if cfg.mode=="local": backup_disk_alert=0 if st.get("continuous",{}).get("state")=="ok" else 1
    else: backup_disk_alert=0 if st.get("present") and st.get("matches") and st.get("continuous",{}).get("state")=="ok" else 1
  except Exception: backup_disk_alert=1
 return JsonResponse({"unread_messages":unread_messages,"pending_reservations":pending_reservations,"unseen_reservations":pending_reservations,"pending_loan_requests":pending_loan_requests,"active_loans":active_loans,"pending_incidents":pending_incidents,"security_red":security_red,"security_yellow":security_yellow,"backup_disk_alert":backup_disk_alert})

@operator_required
def access_control(request):
 now=timezone.now()
 if request.method=="POST":
  action=request.POST.get("action","")
  ip=(request.POST.get("ip_address","") or "").strip()
  if action=="ban":
   duration=(request.POST.get("minutes","60") or "60").strip().lower()
   permanent=duration in {"permanent","permanente","unlimited","ilimitado"}
   minutes=None
   if not permanent:
    try:
     minutes=max(1,min(int(duration),10080))
    except ValueError:
     minutes=60
   reason=(request.POST.get("reason","") or "").strip()[:300]
   if is_protected_local_ip(ip):
    messages.error(request,"Las direcciones locales del servidor (127.0.0.0/8 y ::1) están protegidas y no pueden bloquearse.")
   elif ip==client_ip(request):
    messages.error(request,"No puede bloquear la IP desde la que está administrando la aplicación.")
   else:
    IPBan.objects.filter(ip_address=ip,revoked_at__isnull=True).filter(Q(banned_until__isnull=True)|Q(banned_until__gt=now)).update(revoked_at=now,revoked_by=request.user)
    banned_until=None if permanent else now+timedelta(minutes=minutes)
    IPBan.objects.create(ip_address=ip,banned_by=request.user,banned_until=banned_until,reason=reason)
    AuditLog.objects.create(user=request.user,action="ip_banned",object_type="IP",object_id=ip,details={"minutes":minutes,"permanent":permanent,"reason":reason})
    messages.success(request,f"IP {ip} bloqueada permanentemente." if permanent else f"IP {ip} bloqueada temporalmente.")
  elif action=="unban":
   updated=IPBan.objects.filter(ip_address=ip,revoked_at__isnull=True).filter(Q(banned_until__isnull=True)|Q(banned_until__gt=now)).update(revoked_at=now,revoked_by=request.user)
   if updated:
    AuditLog.objects.create(user=request.user,action="ip_unbanned",object_type="IP",object_id=ip,details={})
    messages.success(request,f"IP {ip} desbloqueada.")
  elif action=="reserve_server_ip":
   try:
    reservation=request_current_ip_reservation(request.user)
    if reservation.status=="applied":
     messages.success(request,f"IP {reservation.ip_address} reservada en DHCP y publicada en DNS correctamente.")
    elif reservation.status=="partial":
     messages.warning(request,f"Solicitud aplicada parcialmente para {reservation.ip_address}. {reservation.message}")
    else:
     messages.warning(request,f"Solicitud registrada para {reservation.ip_address}, pero requiere completar la integración DHCP/DNS. {reservation.message}")
   except Exception as exc:
    AuditLog.objects.create(user=request.user,action="network_ip_reservation_failed",object_type="Network",details={"error":str(exc)[:1000]})
    messages.error(request,f"No se pudo solicitar la reserva de la IP actual: {exc}")
  return redirect("access_control")
 q=(request.GET.get("q","") or "").strip()
 accesses=ServiceAccess.objects.select_related("user").order_by("-last_seen_at")
 if q: accesses=accesses.filter(Q(ip_address__icontains=q)|Q(user__username__icontains=q)|Q(last_path__icontains=q))
 active_bans={b.ip_address:b for b in IPBan.objects.filter(revoked_at__isnull=True).filter(Q(banned_until__isnull=True)|Q(banned_until__gt=now)).select_related("banned_by") if not is_protected_local_ip(b.ip_address)}
 rows=[]
 for access in accesses[:1000]:
  access.active_ban=active_bans.get(access.ip_address); rows.append(access)
 try:
  server_network=current_network()
 except Exception as exc:
  server_network={"error":str(exc)}
 reservations=NetworkReservationRequest.objects.select_related("requested_by").order_by("-requested_at")[:20]
 return render(request,"inventory/access_control.html",{"accesses":rows,"active_bans":list(active_bans.values()),"query":q,"own_ip":client_ip(request),"server_network":server_network,"network_reservations":reservations})

@login_required
def surplus_component(request):
 profile,_=UserProfile.objects.get_or_create(user=request.user)
 if profile.is_guest:
  return HttpResponseForbidden("Los Invitados no pueden registrar componentes sobrantes.")
 locations=_zone_choices(active_only=True)
 active_zone_codes={value for value,_label in locations}
 states=["ÓPTIMO","DAÑADO","KO"]
 if request.method=="POST":
  table=get_object_or_404(InventoryTable,pk=request.POST.get("table_id"),active=True)
  location=(request.POST.get("location") or "").strip()
  state=(request.POST.get("state") or "").strip().upper()
  if location not in active_zone_codes or state not in states:
   messages.error(request,"Seleccione categoría, ubicación y estado válidos.")
  else:
   incident=Incident.objects.create(
    title=f"Componente pendiente de recogida · {table.name}",
    details=f"{request.user.get_username()} comunica un componente sobrante en {_zone_map().get(location,location)}, estado {state}.",
    kind="surplus_pickup",severity="warning",status="pending",
    payload={"user_id":request.user.pk,"username":request.user.get_username(),"table_id":table.pk,"category":table.name,"location":location,"state":state},
   )
   AuditLog.objects.create(user=request.user,action="surplus_component_reported",object_type=table.name,object_id=str(incident.pk),details={"location":location,"location_name":_zone_map().get(location,location),"state":state})
   messages.success(request,"Componente comunicado. Administrador/Gestor recibirá una incidencia para su recogida.")
   return redirect("surplus_component")
 return render(request,"inventory/surplus_component.html",{"tables":InventoryTable.objects.filter(active=True).order_by("name"),"locations":locations,"states":states})

@operator_required
def backup_settings(request):
 from .backup_scheduler import execute_backup
 schedule,_=BackupSchedule.objects.get_or_create(pk=1)
 disk_config,_=BackupDiskConfig.objects.get_or_create(pk=1)
 disk_status={"mounted":False,"source":""}
 if request.user.is_superuser:
  try: disk_status=request_storage_admin({"action":"list_storage","uuid":disk_config.uuid},timeout=4)
  except Exception as exc: disk_status={"mounted":False,"source":"","error":str(exc)}
 if request.method=="POST":
  action=request.POST.get("action","save")
  if action=="configure_local_backup":
   if not request.user.is_superuser: return HttpResponseForbidden("Solo el Gestor puede configurar la protección continua.")
   requested_path=(request.POST.get("local_path") or "").strip()
   try:
    result=request_storage_admin({"action":"configure_local_backup","path":requested_path})
    disk_config.mode="local"; disk_config.local_path=result.get("local_path",""); disk_config.device=""; disk_config.uuid=""; disk_config.filesystem=""; disk_config.mount_point=""; disk_config.last_status="ok"; disk_config.last_error=""; disk_config.updated_by=request.user; disk_config.save()
    schedule.destination=str(Path(disk_config.local_path)/"historico"); Path(schedule.destination).mkdir(parents=True,exist_ok=True); schedule.save(update_fields=["destination"])
    AuditLog.objects.create(user=request.user,action="backup_local_configured",object_type="BackupDiskConfig",object_id=str(disk_config.pk),details={"local_path":disk_config.local_path})
    messages.success(request,f"Protección continua local configurada en {disk_config.local_path}.")
   except Exception as exc:
    disk_config.last_status="error"; disk_config.last_error=str(exc); disk_config.updated_by=request.user; disk_config.save(); messages.error(request,f"No se pudo configurar el directorio local: {exc}")
   return redirect("backup_settings")
  if action=="configure_disk":
   if not request.user.is_superuser: return HttpResponseForbidden("Solo el Gestor puede configurar el disco de backup.")
   device=(request.POST.get("device") or "").strip()
   try:
    result=request_storage_admin({"action":"configure_backup_mount","device":device})
    disk_config.mode="disk"; disk_config.local_path=""; disk_config.device=result.get("device",device); disk_config.uuid=result.get("uuid",""); disk_config.filesystem=result.get("filesystem",""); disk_config.mount_point=result.get("mount_point","/mnt/pulsia-backup"); disk_config.last_status="ok"; disk_config.last_error=""; disk_config.updated_by=request.user; disk_config.save()
    schedule.destination=str(Path(disk_config.mount_point)/"historico"); Path(schedule.destination).mkdir(parents=True,exist_ok=True); schedule.save(update_fields=["destination"])
    AuditLog.objects.create(user=request.user,action="backup_disk_configured",object_type="BackupDiskConfig",object_id=str(disk_config.pk),details={"device":disk_config.device,"uuid":disk_config.uuid,"filesystem":disk_config.filesystem,"mount_point":disk_config.mount_point})
    messages.success(request,f"Disco fijo establecido en {disk_config.mount_point}. La copia continua se ha reiniciado y realizará una copia inicial inmediatamente.")
   except Exception as exc:
    disk_config.last_status="error"; disk_config.last_error=str(exc); disk_config.updated_by=request.user; disk_config.save()
    messages.error(request,f"No se pudo configurar el disco de backup: {exc}")
   return redirect("backup_settings")
  if action=="run_now":
   try:
    target=execute_backup(schedule)
    messages.success(request,f"Copia creada correctamente: {target}")
   except Exception as exc:
    messages.error(request,f"No se pudo crear la copia: {exc}")
   return redirect("backup_settings")
  schedule.enabled=request.POST.get("enabled")=="on"
  destination=(request.POST.get("destination") or "").strip()
  time_text=(request.POST.get("run_time") or "02:00").strip()
  try:
   run_time=datetime.strptime(time_text,"%H:%M").time()
   retention=max(1,min(int(request.POST.get("retention") or 30),1000))
   path=Path(destination).expanduser()
   if not destination: raise ValueError("Indique un destino.")
   path.mkdir(parents=True,exist_ok=True)
   probe=path/".pulsia_write_test"; probe.write_text("ok",encoding="utf-8"); probe.unlink(missing_ok=True)
  except Exception as exc:
   messages.error(request,f"Configuración no válida: {exc}")
   return redirect("backup_settings")
  schedule.destination=str(path); schedule.run_time=run_time; schedule.retention=retention; schedule.updated_by=request.user
  schedule.save()
  AuditLog.objects.create(user=request.user,action="backup_schedule_updated",object_type="BackupSchedule",object_id=str(schedule.pk),details={"enabled":schedule.enabled,"destination":schedule.destination,"run_time":time_text,"retention":retention})
  messages.success(request,"Programación de backups guardada.")
  return redirect("backup_settings")
 now=timezone.localtime()
 next_run=None
 if schedule.enabled:
  candidate=timezone.make_aware(datetime.combine(now.date(),schedule.run_time))
  if candidate<=now: candidate+=timedelta(days=1)
  next_run=candidate
 return render(request,"inventory/backup_settings.html",{"schedule":schedule,"next_run":next_run,"disk_config":disk_config,"disk_status":disk_status,"storage_devices":disk_status.get("devices",[])})

@admin_required
def printing_center(request):
 reprint_form=ExistingLabelReprintForm(prefix="reprint")
 sequence_form=LabelSequenceForm(prefix="sequence")
 if request.method=="POST":
  action=request.POST.get("action","")
  if action=="reprint":
   reprint_form=ExistingLabelReprintForm(request.POST,prefix="reprint")
   if reprint_form.is_valid():
    identifier=reprint_form.identifier(); job=print_identifier(identifier,request.user,2)
    if job.status=="printed": messages.success(request,f"{identifier}: 2 etiquetas enviadas a la Brother QL-700.")
    else: messages.error(request,f"No se pudieron reimprimir las etiquetas de {identifier}: {job.error}")
    return redirect("printing_center")
  elif action=="sequence":
   sequence_form=LabelSequenceForm(request.POST,prefix="sequence")
   if sequence_form.is_valid():
    identifiers=sequence_form.identifiers(); copies=sequence_form.cleaned_data["copies_int"]; results=[print_identifier(identifier,request.user,copies) for identifier in identifiers]
    printed=sum(job.status=="printed" for job in results); failed=len(results)-printed; total_labels=len(identifiers)*copies
    if printed: messages.success(request,f"Secuencia procesada: {printed}/{len(identifiers)} IDs, {copies} copia(s) por ID ({printed*copies} etiquetas impresas de {total_labels} previstas).")
    if failed: messages.error(request,f"{failed} identificador(es) no pudieron imprimirse. Revise la impresora Brother y los últimos trabajos.")
    return redirect("printing_center")
 jobs=LabelPrintJob.objects.select_related("requested_by")[:100]
 return render(request,"inventory/printing.html",{"reprint_form":reprint_form,"sequence_form":sequence_form,"jobs":jobs,"printer_name":os.getenv("LABEL_PRINTER_NAME","Brother QL-700")})

@login_required
@ensure_csrf_cookie
def table_view(request,slug):
 table=get_object_or_404(InventoryTable,slug=slug,active=True); restricted=not request.user.is_staff; fields=[f for f in table.inventory_fields.all() if not restricted or not sensitive_field(f)]; q=request.GET.get("q","").strip(); records=list(table.records.select_related("table").prefetch_related("table__inventory_fields").order_by("internal_id"))
 if restricted: records=[r for r in records if visible_to_normal(r)]
 if q: records=[r for r in records if q.casefold() in (r.internal_id+" "+(public_record_text(r) if restricted else json.dumps(r.data,ensure_ascii=False,default=str))).casefold()]
 page=Paginator(records,100).get_page(request.GET.get("page")); display_rows=[{"record":r,"values":[r.internal_id if f.is_primary else r.data.get(f.key,"") for f in fields]} for r in page.object_list]
 context={"table":table,"fields":fields,"display_rows":display_rows,"page":page,"query":q,"is_admin":request.user.is_staff,"is_operator":request.user.is_superuser,"restricted_user":restricted,"next_page":page.next_page_number() if page.has_next() else None}
 if request.GET.get("fragment")=="1":
  response=render(request,"inventory/_table_rows.html",context); response["X-Next-Page"]=str(context["next_page"] or ""); return response
 return render(request,"inventory/table_view.html",context)

@login_required
def raw_table(request,table=None):
 if table: return redirect("table_view",slug=table)
 first=InventoryTable.objects.filter(active=True).first()
 return redirect("table_view",slug=first.slug) if first else render(request,"inventory/select_table.html",{"tables":[],"title":"No hay tablas todavía"})

@admin_required
def structure(request):
 table_form=InventoryTableForm(request.POST or None,prefix="table"); field_form=InventoryFieldForm(request.POST or None,prefix="field")
 if request.method=="POST":
  form=table_form if "save_table" in request.POST else field_form
  if form.is_valid():
   obj=form.save()
   if isinstance(obj,InventoryTable): obj.created_by=request.user; obj.position=InventoryTable.objects.count(); obj.save(); InventoryField.objects.create(table=obj,name="ID Interno",key="id_interno",position=0,is_primary=True)
   messages.success(request,"Estructura actualizada."); return redirect("structure")
 return render(request,"inventory/structure.html",{"table_form":table_form,"field_form":field_form,"tables":InventoryTable.objects.prefetch_related("inventory_fields")})

@admin_required
def incident_resolve(request,incident_id):
 incident=get_object_or_404(Incident,pk=incident_id)
 if incident.kind!="duplicate_id":
  messages.error(request,"Esta pantalla de propuesta solo está disponible para incidencias de ID duplicado.")
  return redirect("incidents")
 if incident.status in {"resolved","ignored"}:
  messages.warning(request,"La incidencia ya está cerrada.")
  return redirect("incidents")
 table=InventoryTable.objects.filter(pk=incident.payload.get("__inventory_table_pk")).prefetch_related("inventory_fields").first() or InventoryTable.objects.filter(name=incident.source_sheet).prefetch_related("inventory_fields").first()
 if not table:
  messages.error(request,"No se encuentra la tabla de origen de la incidencia.")
  return redirect("incidents")
 primary=table.inventory_fields.filter(is_primary=True).first()
 original_id=(incident.payload.get("__duplicate_internal_id","") or (incident.payload.get(primary.name,"") if primary else "") or incident.title.rsplit(":",1)[-1].strip())
 conflicts=list(InventoryRecord.objects.filter(internal_id__iexact=original_id).select_related("table").prefetch_related("table__inventory_fields").order_by("table__name","pk"))
 existing_record=next((r for r in conflicts if r.table_id==table.pk),None)
 form=DuplicateIncidentResolutionForm(table,request.POST or None,existing_record=existing_record,incoming_payload=incident.payload)
 force_save=request.method=="POST" and request.POST.get("force_save")=="1"
 normal_save=request.method=="POST" and form.is_valid()

 # Guardado forzado controlado: se omiten las validaciones blandas del
 # formulario, pero NUNCA las reglas de identidad. No existe vía para cambiar
 # el internal_id de un registro existente y un alta nueva exige ID no vacío y
 # no duplicado dentro de su tabla.
 if request.method=="POST" and (normal_save or force_save):
  if force_save:
   action=(request.POST.get("resolution_action") or "").strip()
   if action not in {"update_existing","create_new"}:
    messages.error(request,"No se puede forzar el guardado: indique si desea modificar el registro existente o crear uno nuevo.")
    action=""
   if action=="update_existing" and existing_record is None:
    messages.error(request,"No se puede forzar el guardado: no existe un registro de esta tabla que pueda modificarse.")
    action=""
   new_id=(request.POST.get("new_internal_id") or "").strip()
   if action=="create_new":
    if not new_id:
     messages.error(request,"No se puede forzar el guardado: el ID del nuevo registro es obligatorio.")
     action=""
    elif InventoryRecord.objects.filter(table=table,internal_id__iexact=new_id).exists():
     messages.error(request,"No se puede forzar el guardado: ese ID ya existe en esta tabla y no puede suplantarse.")
     action=""
   if action:
    proposed={}
    prefix="existing" if action=="update_existing" else "new"
    for field in table.inventory_fields.filter(is_primary=False):
     proposed[field.key]=request.POST.get(f"{prefix}_{field.key}","")
    note=(request.POST.get("resolution_note") or "").strip() or "Guardado forzado sin nota"
   else:
    proposed={}; note=""
  else:
   action=form.cleaned_data["resolution_action"]
   note=form.cleaned_data["resolution_note"].strip()
   proposed=form.proposed_data("existing" if action=="update_existing" else "new")
   new_id=form.cleaned_data.get("new_internal_id","")

  if action:
   before={}
   print_id=""
   with transaction.atomic():
    if action=="update_existing":
     record=InventoryRecord.objects.select_for_update().get(pk=existing_record.pk)
     immutable_id=record.internal_id
     before={"internal_id":immutable_id,"data":dict(record.data or {}),"status":record.status,"current_sn":record.current_sn,"current_technician":record.current_technician}
     record.data=proposed
     sn_field=table.inventory_fields.filter(is_destination_sn=True).first()
     tech_field=table.inventory_fields.filter(is_technician=True).first()
     record.current_sn=str(proposed.get(sn_field.key,"") or "") if sn_field else record.current_sn
     record.current_technician=str(proposed.get(tech_field.key,"") or "") if tech_field else record.current_technician
     if record.status not in {"scrapped","loaned","reserved"}:
      record.status="assigned" if record.current_sn else "available"
     # Defensa explícita: el ID original vuelve a imponerse antes del save.
     record.internal_id=immutable_id
     record.save()
     RecordMovement.objects.create(record=record,movement_type="correction",reason=f"Resolución {'FORZADA' if force_save else 'manual'} de incidencia #{incident.pk}: {note}",registered_by=request.user)
     result={"resolution_type":"update_existing","forced":force_save,"original_duplicate_id":original_id,"record_pk":record.pk,"record_id":record.internal_id,"before":before,"after":{"data":dict(record.data or {}),"status":record.status,"current_sn":record.current_sn,"current_technician":record.current_technician},"note":note}
     print_id=record.internal_id
     audit_action="duplicate_incident_existing_force_saved" if force_save else "duplicate_incident_existing_updated"
    else:
     sn_field=table.inventory_fields.filter(is_destination_sn=True).first()
     tech_field=table.inventory_fields.filter(is_technician=True).first()
     current_sn=str(proposed.get(sn_field.key,"") or "") if sn_field else ""
     current_technician=str(proposed.get(tech_field.key,"") or "") if tech_field else ""
     record=InventoryRecord.objects.create(table=table,internal_id=new_id,data=proposed,status="assigned" if current_sn else "available",current_sn=current_sn,current_technician=current_technician,created_by=request.user)
     RecordMovement.objects.create(record=record,movement_type="entry",technician_name=current_technician,destination_sn=current_sn,reason=f"Alta {'FORZADA' if force_save else 'propuesta manualmente'} para resolver incidencia #{incident.pk}: {note}",registered_by=request.user)
     result={"resolution_type":"create_new","forced":force_save,"original_duplicate_id":original_id,"new_id":new_id,"resolved_object_pk":record.pk,"after":{"data":dict(record.data or {}),"status":record.status,"current_sn":record.current_sn,"current_technician":record.current_technician},"note":note}
     print_id=new_id
     audit_action="duplicate_incident_new_record_force_saved" if force_save else "duplicate_incident_new_record_created"
    result.update({"resolved_at":timezone.now().isoformat(),"resolved_by":request.user.get_username()})
    payload=dict(incident.payload); payload["resolution"]=result
    incident.payload=payload; incident.status="resolved"; incident.resolved_by=request.user; incident.resolved_at=timezone.now(); incident.save(update_fields=["payload","status","resolved_by","resolved_at"])
    AuditLog.objects.create(user=request.user,action=audit_action,object_type=table.name,object_id=print_id,details={"incident":incident.pk,"forced":force_save,"source_file":incident.source_file,"source_sheet":incident.source_sheet,"source_row":incident.source_row,"resolution":result})
   print_requested=(request.POST.get("print_labels") in {"on","1","true","True"}) if force_save else bool(form.cleaned_data.get("print_labels"))
   if print_requested:
    print_job=print_identifier(print_id,request.user,2)
    payload=dict(incident.payload); resolution=dict(payload.get("resolution",{})); resolution.update({"print_job":print_job.pk,"print_status":print_job.status,"print_error":print_job.error}); payload["resolution"]=resolution; incident.payload=payload; incident.save(update_fields=["payload"])
    if print_job.status=="printed": messages.success(request,f"{'Guardado forzado' if force_save else 'Propuesta aplicada'} e incidencia resuelta. Se imprimieron 2 etiquetas para {print_id}.")
    else: messages.warning(request,f"Registro guardado e incidencia resuelta, pero la impresión falló: {print_job.error}")
   else:
    messages.success(request,"Registro guardado de todos modos e incidencia resuelta." if force_save else "Propuesta aplicada. La incidencia ha quedado resuelta.")
   return redirect("incidents")
 fields=list(table.inventory_fields.all())
 incoming_rows=[]
 for field in fields:
  if field.is_primary:
   incoming=original_id
  else:
   incoming=incident.payload.get(field.name,"")
  incoming_rows.append({"field":field,"incoming":incoming})
 conflict_rows=[]
 for record in conflicts:
  values=[]
  for field in record.table.inventory_fields.all():
   values.append({"name":field.name,"value":record.internal_id if field.is_primary else (record.data or {}).get(field.key,"")})
  conflict_rows.append({"record":record,"values":values})
 return render(request,"inventory/incident_resolve.html",{"incident":incident,"table":table,"original_id":original_id,"conflict_rows":conflict_rows,"existing_record":existing_record,"incoming_rows":incoming_rows,"form":form,"fields":fields})

@admin_required
def incidents_view(request):
 if request.method=="POST":
  incident=get_object_or_404(Incident,pk=request.POST.get("incident_id")); action=request.POST.get("action") or request.POST.get("status")
  if action=="resolve_duplicate":
   return redirect("incident_resolve",incident_id=incident.pk)
  elif action=="reprint_duplicate":
   new_id=(incident.payload.get("resolution") or {}).get("new_id","")
   if not new_id:
    messages.error(request,"La incidencia no tiene un identificador nuevo para reimprimir.")
   else:
    print_job=print_identifier(new_id,request.user,2)
    if print_job.status=="printed": messages.success(request,f"Etiquetas de {new_id} reimpresas por duplicado.")
    else: messages.error(request,f"No se pudieron reimprimir las etiquetas: {print_job.error}")
  elif action=="pickup_collected":
   if incident.kind!="surplus_pickup": return HttpResponseForbidden("Acción no válida.")
   incident.status="resolved"; incident.resolved_by=request.user; incident.resolved_at=timezone.now(); incident.save(update_fields=["status","resolved_by","resolved_at"])
   AuditLog.objects.create(user=request.user,action="surplus_component_collected",object_type="Incident",object_id=str(incident.pk),details=incident.payload)
   messages.success(request,"Alerta retirada: componente marcado como recogido.")
  elif action in {"resolved","ignored","review"}:
   incident.status=action; incident.resolved_by=request.user if action in {"resolved","ignored"} else None; incident.resolved_at=timezone.now() if action in {"resolved","ignored"} else None; incident.save(); messages.success(request,"Incidencia actualizada.")
  return redirect("incidents")
 incidents=Incident.objects.order_by("status","-created_at")
 return render(request,"inventory/incidents.html",{"incidents":incidents})

@operator_required
def import_view(request):
 if request.method=="POST" and request.FILES.get("file"):
  try: job=import_excel(request.FILES["file"],request.user)
  except Exception as exc: messages.error(request,f"No se pudo importar el Excel: {exc}"); return redirect("import_excel")
  messages.success(request,f"{job.rows_imported} objetos importados; {job.rows_incident} incidencias."); return redirect("dashboard")
 return render(request,"inventory/import.html")

@operator_required
def export_view(request):
 data=export_excel(); response=HttpResponse(data,content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"); response["Content-Disposition"]='attachment; filename="inventario_completo.xlsx"'; return response

@operator_required
def database_backup(request):
 db=settings.DATABASES["default"]
 if db["ENGINE"]!="django.db.backends.sqlite3": return HttpResponse("Copia directa disponible únicamente para SQLite.",status=501)
 backup_dir=settings.BASE_DIR/"backups"/"descargas"; backup_dir.mkdir(parents=True,exist_ok=True)
 snapshot=backup_dir/f"inventario_descarga_{timezone.now():%Y%m%d_%H%M%S_%f}.sqlite3"
 create_sqlite_snapshot(Path(db["NAME"]),snapshot)
 return FileResponse(DeleteOnCloseFile(snapshot),as_attachment=True,filename=f"inventario_{timezone.now():%Y%m%d_%H%M%S}.sqlite3")

_DATABASE_RESTORE_LOCK=threading.Lock()

@operator_required
def database_restore(request):
 db=settings.DATABASES["default"]
 if db["ENGINE"]!="django.db.backends.sqlite3": return HttpResponse("La restauración directa está disponible únicamente para SQLite.",status=501)
 if request.method=="POST":
  upload=request.FILES.get("file"); password=request.POST.get("password","")
  if not request.user.check_password(password): messages.error(request,"La contraseña del gestor no es correcta."); return redirect("database_restore")
  if not upload or not upload.name.lower().endswith((".sqlite3",".sqlite",".db")): messages.error(request,"Seleccione una copia SQLite válida."); return redirect("database_restore")
  data_dir=Path(db["NAME"]).parent; backup_dir=settings.BASE_DIR/"backups"; data_dir.mkdir(parents=True,exist_ok=True); backup_dir.mkdir(parents=True,exist_ok=True)
  temporary=None
  try:
   with tempfile.NamedTemporaryFile(prefix="restore_",suffix=".sqlite3",dir=backup_dir,delete=False) as target:
    temporary=Path(target.name)
    for chunk in upload.chunks(): target.write(chunk)
   with temporary.open("rb") as source_file:
    if source_file.read(16)!=b"SQLite format 3\x00": raise ValueError("El archivo no tiene formato SQLite.")
   source=sqlite3.connect(temporary,timeout=10); check=source.execute("PRAGMA quick_check").fetchone()[0]
   required={"auth_user","django_migrations","inventory_inventoryrecord","inventory_inventorytable"}; present={row[0] for row in source.execute("SELECT name FROM sqlite_master WHERE type='table'")}
   if check!="ok" or not required.issubset(present): raise ValueError("La copia está dañada o no pertenece a PULSIA Inventario.")
   if not _DATABASE_RESTORE_LOCK.acquire(blocking=False): raise ValueError("Ya hay una restauración de base de datos en curso.")
   marker=settings.BASE_DIR/"data"/".maintenance"; marker.write_text("restore\n",encoding="utf-8")
   try:
    # Bloquea peticiones nuevas, deja terminar las que ya estaban en curso y cierra las conexiones de este proceso.
    time.sleep(1.0)
    connections.close_all(); database_path=Path(db["NAME"]); safety=backup_dir/f"antes_restaurar_{timezone.now():%Y%m%d_%H%M%S}.sqlite3"
    create_sqlite_snapshot(database_path,safety)
    # Restauración mediante SQLite Backup API; evita sustituir a ciegas un fichero WAL activo.
    destination=sqlite3.connect(database_path,timeout=10)
    try:
     destination.execute("PRAGMA busy_timeout=10000")
     source.backup(destination,pages=256,sleep=0.05)
     destination.execute("DELETE FROM django_session")
     if destination.execute("PRAGMA quick_check").fetchone()[0]!="ok": raise ValueError("La base restaurada no supera quick_check.")
     destination.commit()
    finally:
     destination.close(); source.close(); connections.close_all()
    call_command("migrate",interactive=False,verbosity=0)
   finally:
    try: source.close()
    except Exception: pass
    connections.close_all()
    marker.unlink(missing_ok=True)
    _DATABASE_RESTORE_LOCK.release()
   temporary.unlink(missing_ok=True)
   return render(request,"inventory/operation_complete.html",{"title":"Copia restaurada","message":f"La base de datos se ha restaurado. Se guardó una copia de seguridad previa como {safety.name}.","login_again":True})
  except Exception as exc:
   if temporary: temporary.unlink(missing_ok=True)
   messages.error(request,f"No se pudo restaurar la copia: {exc}")
 return render(request,"inventory/database_restore.html")

@operator_required
def delete_item(request,pk):
 record=get_object_or_404(InventoryRecord.objects.select_related("table"),pk=pk)
 if request.method=="POST" and request.POST.get("confirm")==record.internal_id:
  if record.reservations.exists() or record.loans.exists():
   messages.error(request,"No se puede eliminar físicamente un objeto con histórico de reservas o préstamos. Registre una merma si debe retirarse del uso."); return redirect("dashboard")
  snapshot={"table":record.table.name,"id":record.internal_id}; record.record_movements.all().delete(); record.delete(); AuditLog.objects.create(user=request.user,action="record_deleted",object_type=snapshot["table"],object_id=snapshot["id"]); messages.success(request,"Objeto eliminado físicamente."); return redirect("dashboard")
 return render(request,"inventory/confirm.html",{"title":f"Eliminar {record.internal_id}","warning":f"Se eliminará de la tabla {record.table.name} junto con su historial.","confirm_label":f"Escriba {record.internal_id} para confirmar"})

@operator_required
def truncate_inventory(request):
 ready_key="truncate_ready_at"; now=timezone.now().timestamp()
 if request.method=="GET": request.session[ready_key]=now+15
 ready_at=float(request.session.get(ready_key,now+15)); remaining=max(0,int(ready_at-now+0.999))
 if request.method=="POST":
  if remaining>0: messages.error(request,f"Espere {remaining} segundos antes de confirmar.")
  elif not request.user.check_password(request.POST.get("password","")): messages.error(request,"La contraseña del gestor no es correcta.")
  elif request.POST.get("confirm")!="VACIAR": messages.error(request,"Escriba VACIAR exactamente para confirmar.")
  else:
   counts={"records":InventoryRecord.objects.count(),"movements":RecordMovement.objects.count(),"reservations":Reservation.objects.count(),"loans":Loan.objects.count()}; ReservationView.objects.all().delete(); Reservation.objects.all().delete(); Loan.objects.all().delete(); RecordMovement.objects.all().delete(); InventoryRecord.objects.all().delete(); Incident.objects.all().delete(); ImportJob.objects.all().delete(); AuditLog.objects.create(user=request.user,action="inventory_truncated",object_type="Database",details=counts); request.session.pop(ready_key,None); messages.success(request,"Datos vaciados; estructura y usuarios conservados."); return redirect("dashboard")
 return render(request,"inventory/confirm.html",{"title":"Vaciar inventario","warning":"Se borrarán objetos, movimientos e incidencias. Se conservarán tablas, campos y usuarios.","confirm_label":"Escriba VACIAR para confirmar","require_password":True,"wait_seconds":remaining})

def _stop_process(): os._exit(0)

@operator_required
def stop_service(request):
 if request.method=="POST":
  if not request.user.check_password(request.POST.get("password","")): messages.error(request,"La contraseña del gestor no es correcta."); return redirect("stop_service")
  if request.POST.get("confirm")!="DETENER": messages.error(request,"Escriba DETENER exactamente para confirmar."); return redirect("stop_service")
  AuditLog.objects.create(user=request.user,action="service_stopped",object_type="Server",details={"ip":client_ip(request)})
  threading.Timer(2.0,_stop_process).start()
  return render(request,"inventory/operation_complete.html",{"title":"Servicio detenido","message":"El servidor se está deteniendo. Para volver a usar la aplicación deberá iniciar de nuevo el servicio.","close_window":True})
 return render(request,"inventory/confirm.html",{"title":"Detener el servicio","warning":"Esta acción desconectará inmediatamente a todos los usuarios de la red.","confirm_label":"Introduzca DETENER para continuar","fixed_confirm":"DETENER","require_password":True,"stop_action":True})

@operator_required
def users_panel(request):
 User=get_user_model()
 for account in User.objects.all(): UserProfile.objects.get_or_create(user=account)
 if request.method=="POST":
  action=request.POST.get("action")
  if action=="set_own_password":
   password=request.POST.get("password",""); confirm=request.POST.get("password_confirm","")
   if len(password)<4: messages.error(request,"La contraseña debe tener al menos 4 caracteres.")
   elif password!=confirm: messages.error(request,"Las contraseñas no coinciden.")
   else:
    profile,_=UserProfile.objects.get_or_create(user=request.user); request.user.set_password(password); request.user.save(update_fields=["password"]); profile.bootstrap_token_hash=""; profile.bootstrap_expires_at=None; profile.bootstrap_used_at=timezone.now(); profile.password_reset_requested_at=None; profile.password_reset_authorized_at=None; profile.must_change_password=False; profile.save(); login(request,request.user,backend="django.contrib.auth.backends.ModelBackend"); AuditLog.objects.create(user=request.user,action="gestor_password_set",object_type="User",object_id=str(request.user.pk),details={"ip":client_ip(request)}); messages.success(request,"Contraseña del Gestor establecida correctamente.")
   return redirect("users_panel")
  if action=="create":
   username=request.POST.get("username","").strip().lower(); password=request.POST.get("password","")
   if not username or len(password)<4: messages.error(request,"Indique usuario y contraseña de al menos 4 caracteres.")
   elif User.objects.filter(username__iexact=username).exists() or username in RESERVED_NAMES: messages.error(request,"Ese usuario no está disponible.")
   else: created=User.objects.create_user(username=username,password=password); UserProfile.objects.create(user=created,role="user"); messages.success(request,"Usuario creado correctamente.")
   return redirect("users_panel")
  user=get_object_or_404(User,pk=request.POST.get("user_id"))
  if user.pk==request.user.pk: return HttpResponseForbidden("Utilice el formulario de contraseña del Gestor para modificar su propia cuenta.")
  if user.is_superuser: return HttpResponseForbidden("Las cuentas de máximos permisos están protegidas y no pueden degradarse desde este panel.")
  if action in {"approve_guest","deny_guest"}:
   upgrade=get_object_or_404(AccessUpgradeRequest,user=user,status="pending")
   profile,_=UserProfile.objects.get_or_create(user=user)
   now=timezone.now()
   if action=="approve_guest":
    profile.role="user"; profile.save(update_fields=["role","updated_at"])
    upgrade.status="approved"; upgrade.decided_at=now; upgrade.decided_by=request.user; upgrade.decision_note="Aprobado por el Gestor"; upgrade.save(update_fields=["status","decided_at","decided_by","decision_note"])
    AuditLog.objects.create(user=request.user,action="guest_upgrade_approved",object_type="User",object_id=str(user.pk),details={"username":user.username,"requested_ip":upgrade.requested_ip})
    messages.success(request,f"{user.username} ha sido ascendido a Usuario.")
   else:
    upgrade.status="denied"; upgrade.decided_at=now; upgrade.decided_by=request.user; upgrade.decision_note="Denegado por el Gestor"; upgrade.save(update_fields=["status","decided_at","decided_by","decision_note"])
    user.is_active=False; user.save(update_fields=["is_active"])
    ip=upgrade.requested_ip or profile.created_ip
    if ip and not is_protected_local_ip(ip):
     IPBan.objects.filter(ip_address=ip,revoked_at__isnull=True).filter(Q(banned_until__isnull=True)|Q(banned_until__gt=now)).update(revoked_at=now,revoked_by=request.user)
     IPBan.objects.create(ip_address=ip,banned_by=request.user,banned_until=None,reason=f"Solicitud de ascenso denegada para {user.username}")
    banned_ip=ip if ip and not is_protected_local_ip(ip) else None
    AuditLog.objects.create(user=request.user,action="guest_upgrade_denied_permanent_ban",object_type="User",object_id=str(user.pk),details={"username":user.username,"banned_ip":banned_ip})
    messages.warning(request,f"Solicitud de {user.username} denegada. Cuenta bloqueada permanentemente"+(f" e IP {banned_ip} baneada." if banned_ip else ". La IP local protegida no se ha bloqueado." if ip else "."))
   return redirect("users_panel")
  if action=="reset":
   profile,_=UserProfile.objects.get_or_create(user=user); user.set_unusable_password(); user.is_active=True; user.save(update_fields=["password","is_active"]); profile.password_reset_authorized_at=timezone.now(); profile.password_reset_requested_at=None; profile.must_change_password=False; profile.save(); AuditLog.objects.create(user=request.user,action="password_reset_authorized",object_type="User",object_id=str(user.pk),details={"username":user.username}); messages.success(request,"Contraseña retirada. El usuario ya puede establecer una nueva desde la pantalla de acceso.")
  elif action=="toggle":
   profile,_=UserProfile.objects.get_or_create(user=user)
   user.is_active=not user.is_active; user.save(update_fields=["is_active"])
   if user.is_active:
    profile.archived_at=None; profile.archived_by=None; profile.archived_reason=""; profile.save(update_fields=["archived_at","archived_by","archived_reason","updated_at"])
  elif action=="restore_archived":
   profile,_=UserProfile.objects.get_or_create(user=user); user.is_active=True; user.save(update_fields=["is_active"])
   profile.archived_at=None; profile.archived_by=None; profile.archived_reason=""; profile.save(update_fields=["archived_at","archived_by","archived_reason","updated_at"])
   AuditLog.objects.create(user=request.user,action="user_restored",object_type="User",object_id=str(user.pk),details={"username":user.username})
   messages.success(request,"Usuario reactivado y devuelto a usuarios activos.")
  elif action=="set_role":
   profile,_=UserProfile.objects.get_or_create(user=user)
   target_role=(request.POST.get("target_role") or "").strip().lower()
   if target_role not in {"guest","user","admin"}:
    messages.error(request,"Nivel de permisos no válido.")
    return redirect("users_panel")
   previous_role="admin" if user.is_staff and not user.is_superuser else ("guest" if profile.is_guest else "user")
   if target_role=="guest":
    user.is_staff=False; user.is_superuser=False
    profile.role="guest"
    # Una degradación a Invitado inicia un ciclo de acceso nuevo.
    AccessUpgradeRequest.objects.filter(user=user).delete()
   elif target_role=="user":
    user.is_staff=False; user.is_superuser=False
    profile.role="user"
   else:
    user.is_staff=True; user.is_superuser=False
    profile.role="user"
   user.save(update_fields=["is_staff","is_superuser"])
   profile.save(update_fields=["role","updated_at"])
   AuditLog.objects.create(
    user=request.user,action="user_role_changed",object_type="User",object_id=str(user.pk),
    details={"username":user.username,"from":previous_role,"to":target_role,"ip":client_ip(request)}
   )
   labels={"guest":"Invitado","user":"Usuario","admin":"Administrador"}
   messages.success(request,f"Permisos de {user.username} cambiados de {labels.get(previous_role,previous_role)} a {labels[target_role]}.")
  elif action=="delete":
   profile,_=UserProfile.objects.get_or_create(user=user); now=timezone.now()
   user.is_active=False; user.save(update_fields=["is_active"])
   profile.archived_at=now; profile.archived_by=request.user; profile.archived_reason=(request.POST.get("reason") or "Usuario retirado desde gestión").strip()[:300]
   profile.save(update_fields=["archived_at","archived_by","archived_reason","updated_at"])
   AuditLog.objects.create(user=request.user,action="user_archived",object_type="User",object_id=str(user.pk),details={"username":user.username,"reason":profile.archived_reason})
   messages.success(request,"Usuario movido a Usuarios bloqueados. Se conserva todo su histórico."); return redirect("users_panel")
  return redirect("users_panel")
 pending_access_requests=AccessUpgradeRequest.objects.filter(status="pending").select_related("user").order_by("requested_at")
 all_accounts=User.objects.select_related("inventory_profile").order_by("username")
 active_accounts=[a for a in all_accounts if a.is_active and not a.inventory_profile.archived_at]
 blocked_accounts=[a for a in all_accounts if not a.is_active or a.inventory_profile.archived_at]
 return render(request,"inventory/users.html",{"accounts":active_accounts,"blocked_accounts":blocked_accounts,"current_has_usable_password":request.user.has_usable_password(),"pending_access_requests":pending_access_requests})



@admin_required
def security_center(request):
 policy=get_policy()
 if request.method=="POST":
  action=request.POST.get("action","")
  if action=="review_event":
   event=get_object_or_404(SecurityAccessEvent,pk=request.POST.get("event_id"))
   event.reviewed=True; event.reviewed_at=timezone.now(); event.reviewed_by=request.user; event.resolution="reviewed"
   event.save(update_fields=["reviewed","reviewed_at","reviewed_by","resolution"])
   AuditLog.objects.create(user=request.user,action="security_event_reviewed",object_type="SecurityAccessEvent",object_id=str(event.pk))
   messages.success(request,"Alerta marcada como revisada.")
  elif action in {"unblock_user","keep_blocked","close_session"}:
   if not request.user.is_superuser:
    return HttpResponseForbidden("Solo el Gestor puede ejecutar esta acción de seguridad.")
   if action=="close_session":
    session=get_object_or_404(ActiveSecuritySession,pk=request.POST.get("session_id"))
    close_security_session(session.session_key)
    AuditLog.objects.create(user=request.user,action="security_session_closed",object_type="User",object_id=str(session.user_id),details={"session":session.pk,"ip":session.ip})
    messages.success(request,f"Sesión de {session.user.get_username()} finalizada.")
   else:
    event=get_object_or_404(SecurityAccessEvent,pk=request.POST.get("event_id"))
    target=event.user
    if action=="unblock_user":
     profile,_=UserProfile.objects.get_or_create(user=target)
     target.is_active=True; target.save(update_fields=["is_active"])
     profile.archived_at=None; profile.archived_by=None; profile.archived_reason=""
     profile.save(update_fields=["archived_at","archived_by","archived_reason","updated_at"])
     event.resolution="unblocked"
     AuditLog.objects.create(user=request.user,action="security_user_unblocked",object_type="User",object_id=str(target.pk),details={"event":event.pk})
     messages.success(request,f"{target.get_username()} ha sido desbloqueado.")
    else:
     event.resolution="kept_blocked"
     AuditLog.objects.create(user=request.user,action="security_user_kept_blocked",object_type="User",object_id=str(target.pk),details={"event":event.pk})
     messages.info(request,f"{target.get_username()} permanece bloqueado.")
    event.reviewed=True; event.reviewed_at=timezone.now(); event.reviewed_by=request.user
    event.save(update_fields=["reviewed","reviewed_at","reviewed_by","resolution"])
  return redirect("security_center")

 active_cutoff=timezone.now()-timedelta(minutes=10)
 ActiveSecuritySession.objects.filter(closed=False,last_activity__lt=active_cutoff).update(closed=True,closed_at=timezone.now())
 alerts=SecurityAccessEvent.objects.select_related("user","reviewed_by").order_by("reviewed","-created_at")[:500]
 active_sessions=ActiveSecuritySession.objects.filter(closed=False,last_activity__gte=active_cutoff).select_related("user").order_by("-last_activity")
 red_count=SecurityAccessEvent.objects.filter(reviewed=False,level="RED").count()
 yellow_count=SecurityAccessEvent.objects.filter(reviewed=False,level="YELLOW").count()
 return render(request,"inventory/security_center.html",{
  "alerts":alerts,"active_sessions":active_sessions,"policy":policy,
  "red_count":red_count,"yellow_count":yellow_count,
 })


@admin_required
def security_policy(request):
 policy=get_policy()
 day_choices=[
  ("0","Lunes"),("1","Martes"),("2","Miércoles"),("3","Jueves"),
  ("4","Viernes"),("5","Sábado"),("6","Domingo"),
 ]
 if request.method=="POST":
  enabled=request.POST.get("enabled")=="on"
  days=request.POST.getlist("days")
  start_text=(request.POST.get("start_time") or "").strip()
  end_text=(request.POST.get("end_time") or "").strip()
  errors=[]
  if enabled and not days: errors.append("Seleccione al menos un día permitido.")
  try: start_time=datetime.strptime(start_text,"%H:%M").time()
  except ValueError: start_time=None; errors.append("La hora de inicio no es válida.")
  try: end_time=datetime.strptime(end_text,"%H:%M").time()
  except ValueError: end_time=None; errors.append("La hora de fin no es válida.")
  if start_time and end_time and start_time==end_time: errors.append("La hora de inicio y fin no pueden ser iguales.")
  if errors:
   for error in errors: messages.error(request,error)
  else:
   policy.enabled=enabled; policy.allowed_days=",".join(sorted(set(days))); policy.start_time=start_time; policy.end_time=end_time
   policy.logout_before_end_seconds=60; policy.updated_by=request.user; policy.save()
   AuditLog.objects.create(user=request.user,action="security_policy_updated",object_type="SecurityAccessPolicy",object_id=str(policy.pk),details={
    "enabled":enabled,"allowed_days":policy.allowed_days,"start_time":start_text,"end_time":end_text,"logout_before_end_seconds":60,
   })
   messages.success(request,"Política horaria guardada y activa en la aplicación.")
   return redirect("security_policy")
 allowed=set((policy.allowed_days or "").split(","))
 return render(request,"inventory/security_policy.html",{"policy":policy,"day_choices":day_choices,"allowed_days":allowed,"state":access_window_state(policy)})


@login_required
def security_fingerprint(request):
 if request.method!="POST":
  return JsonResponse({"ok":False},status=405)
 try:
  payload=json.loads(request.body.decode("utf-8") or "{}")
 except Exception:
  payload={}
 try:
  register_client_fingerprint(request.user,request,payload)
 except Exception as exc:
  return JsonResponse({"ok":False,"error":str(exc)[:200]},status=400)
 return JsonResponse({"ok":True})


@login_required
def security_session_state(request):
 if request.user.is_superuser:
  return JsonResponse({"force_logout":False,"gestor":True})
 policy=get_policy()
 state=access_window_state(policy)
 if policy.enabled and (state["forced_logout"] or not state["allowed"]):
  key=request.session.session_key
  if not SecurityAccessEvent.objects.filter(user=request.user,event_type="AUTO_LOGOUT",created_at__gte=timezone.now()-timedelta(minutes=2)).exists():
   SecurityAccessEvent.objects.create(
    user=request.user,level="YELLOW",event_type="AUTO_LOGOUT",
    description="Sesión cerrada automáticamente 60 segundos antes del fin del horario o al quedar fuera de ventana.",
    ip=client_ip(request),current_data={"seconds_until_end":state.get("seconds_until_end")},
   )
  close_security_session(key); logout(request)
  return JsonResponse({"force_logout":True,"reason":"Fin de horario autorizado"},status=401)
 return JsonResponse({
  "force_logout":False,
  "enabled":policy.enabled,
  "seconds_until_end":state.get("seconds_until_end"),
  "warning":bool(policy.enabled and state.get("seconds_until_end") is not None and state["seconds_until_end"]<=300),
 })



def _zone_queryset(active_only=False):
    qs=ProductionZone.objects.order_by("position","name")
    return qs.filter(is_active=True) if active_only else qs

def _zone_choices(active_only=False):
    return [(z.code,z.name) for z in _zone_queryset(active_only)]

def _zone_map():
    values={z.code:z.name for z in _zone_queryset(False)}
    # Compatibilidad de datos históricos anteriores al catálogo único.
    values.update({"Auditoria":"Auditoría","PlanRenove":"Renove","Direccion":"Dirección"})
    return values

def _active_zone_codes():
    return set(ProductionZone.objects.filter(is_active=True).values_list("code",flat=True))

def _production_date(value, default):
    try:
        return datetime.strptime(value, "%Y-%m-%d").date() if value else default
    except (TypeError, ValueError):
        return default


def _production_audit(request, action, entry=None, details=None):
    AuditLog.objects.create(
        user=request.user,
        action=action,
        object_type="ProductionEntry",
        object_id=str(entry.pk) if entry else "",
        details=details or {},
    )


@login_required
def zones_manager(request):
    if not request.user.is_superuser:
        return HttpResponseForbidden("Sólo el Gestor puede modificar las zonas.")
    if request.method=="POST":
        action=(request.POST.get("action") or "").strip()
        if action=="add":
            name=" ".join((request.POST.get("name") or "").strip().split())
            if not name:
                messages.error(request,"Indique el nombre de la zona.")
            elif ProductionZone.objects.filter(name__iexact=name).exists():
                messages.error(request,"Ya existe una zona con ese nombre.")
            else:
                base=slugify(name)[:45] or "zona"; code=base; n=2
                while ProductionZone.objects.filter(code=code).exists():
                    code=f"{base[:40]}-{n}"; n+=1
                position=(ProductionZone.objects.order_by("-position").values_list("position",flat=True).first() or 0)+10
                zone=ProductionZone.objects.create(code=code,name=name,position=position,is_active=True,created_by=request.user)
                AuditLog.objects.create(user=request.user,action="production_zone_created",object_type="ProductionZone",object_id=str(zone.pk),details={"code":zone.code,"name":zone.name})
                messages.success(request,f"Zona creada: {zone.name}.")
        elif action in {"save","toggle"}:
            zone=get_object_or_404(ProductionZone,pk=request.POST.get("zone_id"))
            if action=="toggle":
                zone.is_active=not zone.is_active; zone.save(update_fields=["is_active","updated_at"])
                AuditLog.objects.create(user=request.user,action="production_zone_toggled",object_type="ProductionZone",object_id=str(zone.pk),details={"code":zone.code,"active":zone.is_active})
                messages.success(request,f"Zona {'activada' if zone.is_active else 'desactivada'}: {zone.name}.")
            else:
                name=" ".join((request.POST.get("name") or "").strip().split())
                try: position=max(0,int(request.POST.get("position") or 0))
                except (TypeError,ValueError): position=zone.position
                if not name:
                    messages.error(request,"El nombre de la zona no puede quedar vacío.")
                elif ProductionZone.objects.exclude(pk=zone.pk).filter(name__iexact=name).exists():
                    messages.error(request,"Ya existe otra zona con ese nombre.")
                else:
                    old=zone.name; zone.name=name; zone.position=position; zone.save(update_fields=["name","position","updated_at"])
                    AuditLog.objects.create(user=request.user,action="production_zone_updated",object_type="ProductionZone",object_id=str(zone.pk),details={"code":zone.code,"old_name":old,"name":zone.name,"position":zone.position})
                    messages.success(request,"Zona actualizada.")
        return redirect("zones_manager")
    return render(request,"inventory/zones_manager.html",{"zones":_zone_queryset(False)})

@login_required
def production_board(request):
    """Pizarra tipo bloc de notas del técnico.

    Usuario: sólo puede modificar la franja horaria actual de su propia jornada.
    Administrador: consulta. Gestor/superuser: puede consultar y corregir registros.
    """
    profile, _ = UserProfile.objects.get_or_create(user=request.user)
    if profile.is_guest:
        return redirect("dashboard")

    now = timezone.localtime()
    selected_date = _production_date(request.GET.get("date"), now.date())
    selected_user = request.user
    can_manage = request.user.is_superuser

    if request.user.is_staff or request.user.is_superuser:
        uid = request.GET.get("user")
        if uid:
            selected_user = get_object_or_404(get_user_model(), pk=uid, is_active=True)

    selected_profile, _ = UserProfile.objects.get_or_create(user=selected_user)
    is_own_current_day = selected_user.pk == request.user.pk and selected_date == now.date()
    can_edit_current = is_own_current_day and not request.user.is_staff
    if can_manage and selected_date == now.date():
        can_edit_current = True

    if request.method == "POST":
        action = request.POST.get("action", "add")

        if action in {"save_mysql_source", "test_mysql_source", "import_mysql_models"}:
            if not request.user.is_superuser:
                return HttpResponseForbidden("Sólo el Gestor puede configurar o utilizar la importación automática MySQL.")
            source = ProductionModelMySQLSource.objects.order_by("pk").first()
            if action == "save_mysql_source":
                host = (request.POST.get("mysql_host") or "").strip()
                database = (request.POST.get("mysql_database") or "").strip()
                username = (request.POST.get("mysql_username") or "").strip()
                password = request.POST.get("mysql_password") or ""
                try:
                    port = int(request.POST.get("mysql_port") or 3306)
                except (TypeError, ValueError):
                    port = 0
                if not host or not database or not username:
                    messages.error(request, "Host/IP, base de datos y usuario MySQL son obligatorios.")
                elif port < 1 or port > 65535:
                    messages.error(request, "El puerto MySQL no es válido.")
                elif not password and not (source and source.encrypted_password):
                    messages.error(request, "Debe indicar la contraseña MySQL la primera vez.")
                else:
                    if source is None:
                        source = ProductionModelMySQLSource()
                    source.host = host
                    source.port = port
                    source.database = database
                    source.username = username
                    if password:
                        source.encrypted_password = encrypt_password(password)
                    source.updated_by = request.user
                    source.save()
                    AuditLog.objects.create(user=request.user, action="production_mysql_source_saved", object_type="ProductionModelMySQLSource", object_id=str(source.pk), details={"host": host, "port": port, "database": database, "username": username, "password_changed": bool(password)})
                    messages.success(request, "Configuración MySQL guardada. La contraseña se almacena cifrada.")
                return redirect(f"{reverse('production_board')}?date={selected_date:%Y-%m-%d}&user={selected_user.id}")
            if source is None or not source.encrypted_password:
                messages.error(request, "Configure primero el origen MySQL.")
                return redirect(f"{reverse('production_board')}?date={selected_date:%Y-%m-%d}&user={selected_user.id}")
            try:
                if action == "test_mysql_source":
                    test_source(source)
                    AuditLog.objects.create(user=request.user, action="production_mysql_source_test_ok", object_type="ProductionModelMySQLSource", object_id=str(source.pk), details={"host": source.host, "port": source.port, "database": source.database, "username": source.username})
                    messages.success(request, "Conexión MySQL correcta y tabla Units accesible.")
                else:
                    rows = fetch_models(source)
                    created = existing = excluded = skipped = 0
                    exclusion_names = {n.casefold() for n in ProductionModelExclusion.objects.values_list("name", flat=True)}
                    for manufacturer, model_name in rows:
                        name = normalize_model(manufacturer, model_name)
                        if not name or len(name) > 160:
                            skipped += 1
                            continue
                        if name.casefold() in exclusion_names:
                            excluded += 1
                            continue
                        obj = ProductionModel.objects.filter(name__iexact=name).first()
                        if obj:
                            if obj.is_active:
                                existing += 1
                            else:
                                excluded += 1
                            continue
                        ProductionModel.objects.create(name=name, created_by=request.user, is_active=True)
                        created += 1
                    AuditLog.objects.create(user=request.user, action="production_models_mysql_imported", object_type="ProductionModelMySQLSource", object_id=str(source.pk), details={"received": len(rows), "created": created, "existing": existing, "excluded": excluded, "skipped": skipped, "host": source.host, "database": source.database})
                    messages.success(request, f"Importación completada: {created} nuevos, {existing} ya existentes, {excluded} bloqueados por exclusión y {skipped} omitidos.")
            except Exception as exc:
                AuditLog.objects.create(user=request.user, action="production_mysql_source_error", object_type="ProductionModelMySQLSource", object_id=str(source.pk), details={"host": source.host, "database": source.database, "error": str(exc)[:500]})
                messages.error(request, f"No se pudo completar la operación MySQL: {exc}")
            return redirect(f"{reverse('production_board')}?date={selected_date:%Y-%m-%d}&user={selected_user.id}")

        if action == "add_model":
            if not (request.user.is_staff or request.user.is_superuser):
                return HttpResponseForbidden("Sólo Gestor o Administradores pueden ampliar el catálogo de modelos.")
            new_name = " ".join((request.POST.get("new_model_name") or "").strip().split()).upper()
            if not new_name:
                messages.error(request, "Indique el nuevo modelo.")
            elif len(new_name) > 160:
                messages.error(request, "El modelo no puede superar 160 caracteres.")
            elif ProductionModelExclusion.objects.filter(name__iexact=new_name).exists():
                messages.warning(request, "Ese modelo está en la lista de exclusiones. Rehabilítelo antes de añadirlo.")
            else:
                model = ProductionModel.objects.filter(name__iexact=new_name).first()
                if model and model.is_active:
                    messages.warning(request, "Ese modelo ya existe en el catálogo.")
                elif model:
                    model.is_active=True; model.save(update_fields=["is_active"])
                    AuditLog.objects.create(user=request.user, action="production_model_reactivated", object_type="ProductionModel", object_id=str(model.pk), details={"name": model.name})
                    messages.success(request, f"Modelo reactivado: {model.name}")
                else:
                    model = ProductionModel.objects.create(name=new_name, created_by=request.user)
                    AuditLog.objects.create(user=request.user, action="production_model_created", object_type="ProductionModel", object_id=str(model.pk), details={"name": model.name})
                    messages.success(request, f"Modelo añadido al catálogo: {model.name}")
            return redirect(f"{reverse('production_board')}?date={selected_date:%Y-%m-%d}&user={selected_user.id}")

        if action == "exclude_model":
            if not (request.user.is_staff or request.user.is_superuser):
                return HttpResponseForbidden("Sólo Gestor o Administradores pueden excluir modelos.")
            model = get_object_or_404(ProductionModel, pk=request.POST.get("model_id"))
            model.is_active=False; model.save(update_fields=["is_active"])
            exclusion, _ = ProductionModelExclusion.objects.get_or_create(name=model.name, defaults={"excluded_by": request.user, "reason": "Retirado del catálogo por Gestor/Administrador"})
            AuditLog.objects.create(user=request.user, action="production_model_excluded", object_type="ProductionModel", object_id=str(model.pk), details={"name": model.name, "exclusion_id": exclusion.pk})
            messages.success(request, f"Modelo retirado y añadido a exclusiones: {model.name}")
            return redirect(f"{reverse('production_board')}?date={selected_date:%Y-%m-%d}&user={selected_user.id}")

        if action == "restore_model_exclusion":
            if not (request.user.is_staff or request.user.is_superuser):
                return HttpResponseForbidden("Sólo Gestor o Administradores pueden rehabilitar modelos.")
            exclusion = get_object_or_404(ProductionModelExclusion, pk=request.POST.get("exclusion_id"))
            name = exclusion.name
            model = ProductionModel.objects.filter(name__iexact=name).first()
            if model:
                model.is_active=True; model.save(update_fields=["is_active"])
            else:
                model = ProductionModel.objects.create(name=name, created_by=request.user, is_active=True)
            exclusion.delete()
            AuditLog.objects.create(user=request.user, action="production_model_exclusion_restored", object_type="ProductionModel", object_id=str(model.pk), details={"name": name})
            messages.success(request, f"Modelo rehabilitado: {name}")
            return redirect(f"{reverse('production_board')}?date={selected_date:%Y-%m-%d}&user={selected_user.id}")

        if action == "add_processor":
            if not request.user.is_superuser:
                return HttpResponseForbidden("Sólo el Gestor puede ampliar el catálogo de procesadores.")
            new_name = " ".join((request.POST.get("new_processor_name") or "").strip().split()).upper()
            if not new_name:
                messages.error(request, "Indique el nuevo procesador.")
            elif len(new_name) > 160:
                messages.error(request, "El procesador no puede superar 160 caracteres.")
            elif ProductionProcessor.objects.filter(name__iexact=new_name).exists():
                messages.warning(request, "Ese procesador ya existe en el catálogo.")
            else:
                processor = ProductionProcessor.objects.create(name=new_name, created_by=request.user)
                AuditLog.objects.create(user=request.user, action="production_processor_created", object_type="ProductionProcessor", object_id=str(processor.pk), details={"name": processor.name})
                messages.success(request, f"Procesador añadido al catálogo: {processor.name}")
            return redirect(f"{reverse('production_board')}?date={selected_date:%Y-%m-%d}&user={selected_user.id}")

        entry = None
        entry_id = request.POST.get("entry_id")
        if entry_id:
            entry = get_object_or_404(ProductionEntry, pk=entry_id)

        # Usuario normal: únicamente sus registros de la hora actual. Gestor: corrección explícita.
        if entry:
            owner_current = entry.user_id == request.user.id and entry.date == now.date() and entry.hour == now.hour and not request.user.is_staff
            gestor_allowed = request.user.is_superuser
            if not (owner_current or gestor_allowed):
                return HttpResponseForbidden("La franja está cerrada o no tiene permiso para modificarla.")
            if action in {"inc", "dec", "set"}:
                before = entry.quantity
                if action == "inc":
                    entry.quantity += 1
                elif action == "dec":
                    entry.quantity = max(0, entry.quantity - 1)
                else:
                    try:
                        entry.quantity = max(0, int(request.POST.get("quantity", entry.quantity)))
                    except (TypeError, ValueError):
                        messages.error(request, "La cantidad indicada no es válida.")
                        return redirect(f"{reverse('production_board')}?date={entry.date:%Y-%m-%d}&user={entry.user_id}")
                if entry.quantity == 0:
                    _production_audit(request, "production_entry_deleted", entry, {"before": before})
                    entry.delete()
                else:
                    entry.save(update_fields=["quantity"])
                    _production_audit(request, "production_entry_quantity_changed", entry, {"before": before, "after": entry.quantity})
                return redirect(f"{reverse('production_board')}?date={entry.date:%Y-%m-%d}&user={entry.user_id}")

        if action == "add":
            if not can_edit_current:
                return HttpResponseForbidden("Sólo puede registrar producción en la franja horaria actual.")
            model_id = request.POST.get("production_model", "").strip()
            processor_id = request.POST.get("processor", "").strip()
            origin_zone = request.POST.get("origin_zone", "").strip()
            zone = request.POST.get("zone", "").strip()
            production_model = ProductionModel.objects.filter(pk=model_id, is_active=True).first() if model_id.isdigit() else None
            processor = ProductionProcessor.objects.filter(pk=processor_id, is_active=True).first() if processor_id.isdigit() else None
            try:
                ram_gb = int(request.POST.get("ram_gb", "0") or 0)
                disk_gb = int(request.POST.get("disk_gb", "0") or 0)
                qty = int(request.POST.get("quantity", "1") or 1)
            except (TypeError, ValueError):
                ram_gb = disk_gb = qty = 0
            if not production_model:
                messages.error(request, "Seleccione un modelo válido del catálogo.")
            elif ram_gb < 1:
                messages.error(request, "Indique una cantidad válida de RAM en GB.")
            elif disk_gb < 1:
                messages.error(request, "Indique una capacidad válida de disco en GB.")
            elif not processor:
                messages.error(request, "Seleccione un procesador válido del catálogo.")
            elif origin_zone not in _active_zone_codes():
                messages.error(request, "Seleccione una zona de origen válida. No se puede registrar producción sin origen.")
            elif zone not in _active_zone_codes():
                messages.error(request, "Seleccione una zona de destino válida. No se puede registrar producción sin destino.")
            elif qty < 1:
                messages.error(request, "La cantidad debe ser al menos 1.")
            else:
                target_user = selected_user if can_manage else request.user
                entry = ProductionEntry.objects.create(
                    user=target_user,
                    date=selected_date if can_manage else now.date(),
                    hour=now.hour,
                    model_name=production_model.name,
                    production_model=production_model,
                    ram_gb=ram_gb,
                    disk_gb=disk_gb,
                    processor=processor,
                    processor_name=processor.name,
                    origin_zone=origin_zone,
                    zone=zone,
                    quantity=qty,
                )
                _production_audit(request, "production_entry_created", entry, {"model": entry.model_name, "ram_gb": ram_gb, "disk_gb": disk_gb, "processor": processor.name, "origin_zone": entry.origin_zone, "destination_zone": entry.zone, "quantity": qty})
                messages.success(request, "Producción añadida a la pizarra.")
            return redirect(f"{reverse('production_board')}?date={selected_date:%Y-%m-%d}&user={selected_user.id}")

    entries = list(
        ProductionEntry.objects.filter(user=selected_user, date=selected_date)
        .order_by("hour", "created_at", "pk")
    )
    total_day = sum(e.quantity for e in entries)
    hours = []
    for hour in range(24):
        hour_entries = [e for e in entries if e.hour == hour]
        if not hour_entries and not (selected_date == now.date() and hour == now.hour):
            continue
        hours.append({
            "hour": hour,
            "label": f"{hour:02d}:00 – {(hour + 1) % 24:02d}:00",
            "entries": hour_entries,
            "total": sum(e.quantity for e in hour_entries),
            "editable": (selected_date == now.date() and hour == now.hour and ((selected_user.pk == request.user.pk and not request.user.is_staff) or can_manage)),
            "closed": selected_date < now.date() or (selected_date == now.date() and hour < now.hour),
        })
    users = get_user_model().objects.filter(is_active=True).order_by("first_name", "last_name", "username") if request.user.is_staff else []
    production_models = ProductionModel.objects.filter(is_active=True).order_by("name")
    all_production_models = ProductionModel.objects.order_by("name") if request.user.is_staff else []
    model_exclusions = ProductionModelExclusion.objects.order_by("name") if request.user.is_staff else []
    processors = ProductionProcessor.objects.filter(is_active=True).order_by("name")
    mysql_source = ProductionModelMySQLSource.objects.order_by("pk").first() if request.user.is_superuser else None
    return render(request, "inventory/production_board.html", {
        "hours": hours,
        "total_day": total_day,
        "current_hour": now.hour,
        "selected_date": selected_date,
        "selected_user": selected_user,
        "selected_profile": selected_profile,
        "users": users,
        "zones": _zone_choices(active_only=True),
        "production_models": production_models,
        "all_production_models": all_production_models,
        "model_exclusions": model_exclusions,
        "processors": processors,
        "can_add_processor": request.user.is_superuser,
        "can_add_model": request.user.is_staff or request.user.is_superuser,
        "can_auto_import_models": request.user.is_superuser,
        "mysql_source": mysql_source,
        "can_edit_current": can_edit_current,
        "can_manage": can_manage,
        "today": now.date(),
    })


@admin_required
def production_current(request):
    """Producción del día en curso para Administrador y Gestor.

    Vista estrictamente de consulta. Los usuarios normales no pueden acceder
    aunque conozcan la URL.
    """
    now = timezone.localtime()
    today = now.date()
    current_hour = now.hour
    entries = (
        ProductionEntry.objects
        .filter(date=today)
        .select_related("user", "production_model", "processor")
    )

    total_day = entries.aggregate(total=Sum("quantity"))["total"] or 0
    current_entries = entries.filter(hour=current_hour)
    total_current_hour = current_entries.aggregate(total=Sum("quantity"))["total"] or 0
    active_users = entries.values("user_id").distinct().count()

    zone_rows = list(
        entries.values("zone")
        .annotate(total=Sum("quantity"))
        .order_by("zone")
    )
    current_zone_totals = {
        row["zone"]: row["total"]
        for row in current_entries.values("zone").annotate(total=Sum("quantity"))
    }
    for row in zone_rows:
        row["label"] = _zone_map().get(row["zone"], row["zone"])
        row["current_total"] = current_zone_totals.get(row["zone"], 0)

    user_rows = list(
        entries.values("user_id", "user__username", "user__first_name", "user__last_name")
        .annotate(total=Sum("quantity"))
        .order_by("-total", "user__username")
    )
    current_user_totals = {
        row["user_id"]: row["total"]
        for row in current_entries.values("user_id").annotate(total=Sum("quantity"))
    }
    for row in user_rows:
        full_name = f"{row['user__first_name']} {row['user__last_name']}".strip()
        row["display_name"] = full_name or row["user__username"]
        row["current_total"] = current_user_totals.get(row["user_id"], 0)

    recent_entries = entries.order_by("-created_at")[:100]
    return render(request, "inventory/production_current.html", {
        "today": today,
        "now": now,
        "current_hour": current_hour,
        "total_day": total_day,
        "total_current_hour": total_current_hour,
        "active_users": active_users,
        "zone_rows": zone_rows,
        "user_rows": user_rows,
        "recent_entries": recent_entries,
    })


@admin_required
def production_reports(request):
    now = timezone.localtime()
    end = _production_date(request.GET.get("end"), now.date())
    start = _production_date(request.GET.get("start"), end)
    if start > end:
        start, end = end, start
    qs = ProductionEntry.objects.select_related("user","processor").filter(date__range=(start, end))
    user_id = request.GET.get("user", "").strip()
    zone = request.GET.get("zone", "").strip()
    origin_zone = request.GET.get("origin_zone", "").strip()
    hour_text = request.GET.get("hour", "").strip()
    selected_hour = ""
    if user_id:
        qs = qs.filter(user_id=user_id)
    if zone in _zone_map():
        qs = qs.filter(zone=zone)
    elif zone:
        zone = ""
    if origin_zone in _zone_map():
        qs = qs.filter(origin_zone=origin_zone)
    elif origin_zone:
        origin_zone = ""
    if hour_text:
        try:
            hour_value=int(hour_text)
            if 0 <= hour_value <= 23:
                qs=qs.filter(hour=hour_value); selected_hour=str(hour_value)
        except (TypeError,ValueError):
            pass

    user_summary = list(qs.values("user_id", "user__username", "user__first_name", "user__last_name").annotate(total=Sum("quantity"), lines=Count("id")).order_by("user__first_name", "user__last_name", "user__username"))
    zone_summary = list(qs.values("zone").annotate(total=Sum("quantity"), lines=Count("id")).order_by("zone"))
    for row in zone_summary:
        row["zone_name"] = _zone_map().get(row["zone"], row["zone"])
    detail = list(qs.order_by("-date", "hour", "user__username", "zone", "model_name"))
    grand_total = sum(e.quantity for e in detail)
    users = get_user_model().objects.filter(is_active=True).order_by("first_name", "last_name", "username")
    return render(request, "inventory/production_reports.html", {
        "start": start, "end": end, "selected_user_id": user_id, "selected_zone": zone,
        "selected_origin_zone": origin_zone, "selected_hour": selected_hour, "hours_filter": range(24),
        "users": users, "zones": _zone_choices(active_only=False), "user_summary": user_summary,
        "zone_summary": zone_summary, "detail": detail, "grand_total": grand_total,
    })


@admin_required
def production_export(request):
    from openpyxl import Workbook
    from openpyxl.styles import Font
    now = timezone.localtime()
    end = _production_date(request.GET.get("end"), now.date())
    start = _production_date(request.GET.get("start"), end)
    if start > end:
        start, end = end, start
    qs = ProductionEntry.objects.select_related("user","processor").filter(date__range=(start, end))
    user_id = request.GET.get("user", "").strip()
    zone = request.GET.get("zone", "").strip()
    origin_zone = request.GET.get("origin_zone", "").strip()
    hour_text = request.GET.get("hour", "").strip()
    if user_id:
        qs = qs.filter(user_id=user_id)
    if zone in _zone_map():
        qs = qs.filter(zone=zone)
    if origin_zone in _zone_map():
        qs = qs.filter(origin_zone=origin_zone)
    if hour_text:
        try:
            hour_value=int(hour_text)
            if 0 <= hour_value <= 23: qs=qs.filter(hour=hour_value)
        except (TypeError,ValueError): pass

    wb = Workbook()
    ws = wb.active
    ws.title = "Producción"
    headers = ["Fecha", "Hora", "Usuario", "Zona origen", "Zona destino", "Modelo", "RAM_GB", "Disco_GB", "Procesador", "Cantidad"]
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True)
    for entry in qs.order_by("date", "hour", "user__username", "zone", "model_name"):
        full_name = entry.user.get_full_name().strip() or entry.user.username
        ws.append([entry.date.strftime("%d/%m/%Y"), f"{entry.hour:02d}:00-{(entry.hour+1)%24:02d}:00", full_name, _zone_map().get(entry.origin_zone, entry.origin_zone or "—"), entry.get_zone_display(), entry.model_name, entry.ram_gb or "", entry.disk_gb or "", entry.processor_name or (entry.processor.name if entry.processor_id else ""), entry.quantity])
    ws.append([])
    ws.append(["TOTAL", "", "", "", "", "", "", "", "", qs.aggregate(total=Sum("quantity"))["total"] or 0])
    for width, col in zip((14,16,24,18,18,30,10,12,28,12), "ABCDEFGHIJ"):
        ws.column_dimensions[col].width = width

    by_user = wb.create_sheet("Resumen usuarios")
    by_user.append(["Usuario", "Unidades"])
    for cell in by_user[1]: cell.font = Font(bold=True)
    for row in qs.values("user__username", "user__first_name", "user__last_name").annotate(total=Sum("quantity")).order_by("user__username"):
        name = (f"{row['user__first_name']} {row['user__last_name']}".strip() or row["user__username"])
        by_user.append([name, row["total"]])

    by_zone = wb.create_sheet("Resumen zonas")
    by_zone.append(["Zona", "Unidades"])
    for cell in by_zone[1]: cell.font = Font(bold=True)
    for row in qs.values("zone").annotate(total=Sum("quantity")).order_by("zone"):
        by_zone.append([_zone_map().get(row["zone"], row["zone"]), row["total"]])

    import io
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return FileResponse(output, as_attachment=True, filename=f"produccion_{start:%Y%m%d}_{end:%Y%m%d}.xlsx")
